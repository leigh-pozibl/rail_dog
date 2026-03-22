"""
Reporting utilities for generating Excel and Power BI compatible outputs from rail_dog database.

This module provides functions to:
1. Export aggregated rail data to Excel workbooks with tabs per line_code
2. Generate Power BI compatible exports with geospatial data
3. Create unified datasets for visualization
"""
import logging
import os
from datetime import datetime
from typing import Optional, List, Dict
import pandas as pd
import geopandas as gpd
from sqlalchemy import Engine
from sqlalchemy.orm import Session
from sqlmodel import select

from rail_dog.configs.schema import RailSegment, AggAsset, AggGBFI, AggTQI
from rail_dog.utils.db_utils import get_table_data, query_agg_tsr


def generate_excel_report(
    engine: Engine,
    output_path: str,
    line_codes: Optional[List[str]] = None
) -> str:
    """
    Generate an Excel workbook with one tab per line_code, joining all aggregated tables.

    Args:
        engine: SQLAlchemy database engine
        output_path: Path to save the Excel file
        line_codes: Optional list of line codes to include (default: all)

    Returns:
        Path to the generated Excel file
    """
    logging.info("Generating Excel report...")
    
    output_file = os.path.join(output_path, "fmg_rail_report.xlsx")

    # If no line codes specified, get all unique line codes from segments
    if line_codes is None:
        segments_df = get_table_data(engine, table_name="rail_segments")
        line_codes = sorted(segments_df['line_code'].unique())
        
    logging.info(f"Found line codes: {line_codes}")

    # Load all data tables once
    logging.info("Loading data from database...")
    segments_df = get_table_data(engine, table_name="rail_segments")
    agg_asset_df = get_table_data(engine, table_name="agg_assets")
    agg_gbfi_df = get_table_data(engine, table_name="agg_gbfi")
    agg_tsr_df = query_agg_tsr(engine)
    agg_tqi_df = get_table_data(engine, table_name="agg_tqi")
    agg_dtr_df = get_table_data(engine, table_name="agg_dtr")
    agg_tg_df = get_table_data(engine, query="SELECT chainage_id, avg_speed FROM agg_tg")

    # Remove geometry and metadata columns from segments for Excel export
    segments_export = segments_df.drop(columns=['geometry', 'created_at', 'id'], errors='ignore')

    # Map avg_speed from agg_tg onto segments
    if not agg_tg_df.empty and 'avg_speed' in agg_tg_df.columns:
        speed_map = agg_tg_df.set_index('chainage_id')['avg_speed']
        segments_export['speed'] = segments_export['chainage_id'].map(speed_map)

    # Remove id and created_at from agg tables
    agg_asset_export = agg_asset_df.drop(columns=['id', 'created_at'], errors='ignore')
    agg_gbfi_export = agg_gbfi_df.drop(columns=['id', 'collection_date', 'created_at'], errors='ignore')
    agg_tsr_export = agg_tsr_df.drop(columns=['id', 'created_at'], errors='ignore')
    agg_tqi_export = agg_tqi_df.drop(columns=['id', 'collection_date', 'created_at'], errors='ignore')
    agg_dtr_export = agg_dtr_df.drop(columns=['id', 'collection_date', 'created_at'], errors='ignore')
    # agg_tg_export = agg_tg_df[['chainage_id', 'avg_speed']].copy() if 'avg_speed' in agg_tg_df.columns else agg_tg_df[['chainage_id']].copy()

    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for line_code in line_codes:
            logging.info(f"Processing line_code: {line_code}")

            # Filter segments for this line_code
            line_segments = segments_export[segments_export['line_code'] == line_code].copy()

            if line_segments.empty:
                logging.warning(f"No segments found for line_code: {line_code}")
                continue

            # Track column sources for header row
            segment_cols = list(line_segments.columns)

            # Merge all aggregated data on chainage_id
            merged = line_segments.copy()

            # Left join each agg table and track new columns
            asset_cols = []
            if not agg_asset_export.empty:
                before_cols = set(merged.columns)
                merged = merged.merge(
                    agg_asset_export,
                    on='chainage_id',
                    how='left',
                    suffixes=('', '_asset')
                )
                asset_cols = [col for col in merged.columns if col not in before_cols]

            gbfi_cols = []
            if not agg_gbfi_export.empty:
                before_cols = set(merged.columns)
                merged = merged.merge(
                    agg_gbfi_export,
                    on='chainage_id',
                    how='left',
                    suffixes=('', '_gbfi')
                )
                gbfi_cols = [col for col in merged.columns if col not in before_cols]

            tsr_cols = []
            if not agg_tsr_export.empty:
                before_cols = set(merged.columns)
                merged = merged.merge(
                    agg_tsr_export,
                    on='chainage_id',
                    how='left',
                    suffixes=('', '_tsr')
                )
                tsr_cols = [col for col in merged.columns if col not in before_cols]

            tqi_cols = []
            if not agg_tqi_export.empty:
                before_cols = set(merged.columns)
                merged = merged.merge(
                    agg_tqi_export,
                    on='chainage_id',
                    how='left',
                    suffixes=('', '_tqi')
                )
                tqi_cols = [col for col in merged.columns if col not in before_cols]
            
            dtr_cols = []
            if not agg_dtr_export.empty:
                before_cols = set(merged.columns)
                merged = merged.merge(
                    agg_dtr_export,
                    on='chainage_id',
                    how='left',
                    suffixes=('', '_dtr')
                )
                dtr_cols = [col for col in merged.columns if col not in before_cols]

            # Sort by chainage_start_km
            merged = merged.sort_values('chainage_start_km')

            # Write to Excel tab (sheet name limited to 31 chars)
            sheet_name = f"{line_code}"[:31]
            merged.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

            # Access the worksheet to add merged header row
            worksheet = writer.sheets[sheet_name]

            # Build column group mapping
            col_groups = []

            # Determine source for each column
            start_col = 1  # Excel columns are 1-indexed

            if segment_cols:
                end_col = start_col + len(segment_cols) - 1
                col_groups.append(('Rail Segments', start_col, end_col))
                start_col = end_col + 1

            if asset_cols:
                end_col = start_col + len(asset_cols) - 1
                col_groups.append(('Assets', start_col, end_col))
                start_col = end_col + 1

            if gbfi_cols:
                end_col = start_col + len(gbfi_cols) - 1
                col_groups.append(('Ground Ballast Fouling Index (GBFI)', start_col, end_col))
                start_col = end_col + 1

            if tsr_cols:
                end_col = start_col + len(tsr_cols) - 1
                col_groups.append(('Temporary Speed Restriction (TSR)', start_col, end_col))
                start_col = end_col + 1

            if tqi_cols:
                end_col = start_col + len(tqi_cols) - 1
                col_groups.append(('Track Quality Indicator (TQI)', start_col, end_col))
                start_col = end_col + 1
                
            if dtr_cols:
                end_col = start_col + len(dtr_cols) - 1
                col_groups.append(('Dynamic Track Force (DTR)', start_col, end_col))

            # Insert merged cells in first row
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter

            for source_name, start_idx, end_idx in col_groups:
                start_letter = get_column_letter(start_idx)
                end_letter = get_column_letter(end_idx)

                # Merge cells
                if start_idx == end_idx:
                    cell_ref = f"{start_letter}1"
                else:
                    cell_ref = f"{start_letter}1:{end_letter}1"
                    worksheet.merge_cells(cell_ref)

                # Write source label
                cell = worksheet[f"{start_letter}1"]
                cell.value = source_name
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Add background color based on source
                colors = {
                    'Rail Segments': 'DDEBF7',  # Light blue
                    'Assets': 'E2EFDA',    # Light green
                    'Ground Ballast Fouling Index (GBFI)': 'FCE4D6',     # Light orange
                    'Temporary Speed Restriction (TSR)': 'FFF2CC',      # Light yellow
                    'Track Quality Indicator (TQI)': 'F4B084',      # Light coral
                    'Dynamic Track Force (DTR)': 'E4DFEC',      # Light purple
                }
                if source_name in colors:
                    cell.fill = PatternFill(start_color=colors[source_name],
                                           end_color=colors[source_name],
                                           fill_type='solid')

            # Set row height for header
            worksheet.row_dimensions[1].height = 20

            logging.info(f"  Written {len(merged)} rows to sheet '{sheet_name}'")

    logging.info(f"Excel report saved to: {output_path}")
    return output_path


def generate_powerbi_export(
    engine: Engine,
    output_dir: str,
    format: str = 'geojson'
) -> Dict[str, str]:
    """
    Generate Power BI compatible exports with geospatial data.

    Creates separate files:
    - rail_segments.geojson: Segment geometries with all attributes
    - agg_assets.csv: Asset counts by chainage_id
    - agg_gbfi.csv: GBFI metrics by chainage_id
    - agg_tsr.csv: TSR data by chainage_id
    - agg_tqi.csv: TQI data by chainage_id
    - unified_rail_data.geojson: All data joined (for direct use in Power BI)

    Args:
        engine: SQLAlchemy database engine
        output_dir: Directory to save output files
        format: Output format for geospatial data ('geojson', 'shapefile', 'geoparquet')

    Returns:
        Dictionary mapping file type to file path
    """
    logging.info("Generating Power BI compatible exports...")
    os.makedirs(output_dir, exist_ok=True)

    output_files = {}

    # Load segments with geometry
    logging.info("Loading rail segments...")
    query = "SELECT * FROM rail_segments"
    segments_gdf = gpd.read_postgis(
        query,
        engine,
        geom_col='geometry',
        crs='EPSG:4326'
    )

    # Load aggregated tables
    logging.info("Loading aggregated data...")
    agg_asset_df = get_table_data(engine, table_name="agg_assets")
    agg_gbfi_df = get_table_data(engine, table_name="agg_gbfi")
    agg_tsr_df = query_agg_tsr(engine)
    agg_tqi_df = get_table_data(engine, table_name="agg_tqi")

    # Export individual CSV files (without geometry, for Power BI relationships)
    logging.info("Exporting individual CSV files...")

    csv_path = os.path.join(output_dir, "agg_assets.csv")
    agg_asset_df.drop(columns=['id', 'created_at'], errors='ignore').to_csv(csv_path, index=False)
    output_files['agg_assets'] = csv_path

    csv_path = os.path.join(output_dir, "agg_gbfi.csv")
    agg_gbfi_df.drop(columns=['id', 'created_at'], errors='ignore').to_csv(csv_path, index=False)
    output_files['agg_gbfi'] = csv_path

    csv_path = os.path.join(output_dir, "agg_tsr.csv")
    agg_tsr_df.drop(columns=['id', 'created_at'], errors='ignore').to_csv(csv_path, index=False)
    output_files['agg_tsr'] = csv_path

    csv_path = os.path.join(output_dir, "agg_tqi.csv")
    agg_tqi_df.drop(columns=['id', 'created_at'], errors='ignore').to_csv(csv_path, index=False)
    output_files['agg_tqi'] = csv_path

    # Export segments with geometry
    logging.info("Exporting rail segments with geometry...")
    segments_export = segments_gdf.drop(columns=['id', 'created_at'], errors='ignore')

    if format == 'geojson':
        geo_path = os.path.join(output_dir, "rail_segments.geojson")
        segments_export.to_file(geo_path, driver='GeoJSON')
    elif format == 'shapefile':
        geo_path = os.path.join(output_dir, "rail_segments.shp")
        segments_export.to_file(geo_path)
    elif format == 'geoparquet':
        geo_path = os.path.join(output_dir, "rail_segments.parquet")
        segments_export.to_parquet(geo_path)

    output_files['rail_segments'] = geo_path

    # Create unified dataset (join all on chainage_id)
    logging.info("Creating unified dataset with all metrics...")
    unified = segments_gdf.copy()

    # Drop metadata columns
    unified = unified.drop(columns=['id', 'created_at'], errors='ignore')

    # Left join all agg tables
    if not agg_asset_df.empty:
        unified = unified.merge(
            agg_asset_df.drop(columns=['id', 'created_at'], errors='ignore'),
            on='chainage_id',
            how='left',
            suffixes=('', '_asset')
        )

    if not agg_gbfi_df.empty:
        unified = unified.merge(
            agg_gbfi_df.drop(columns=['id', 'created_at'], errors='ignore'),
            on='chainage_id',
            how='left',
            suffixes=('', '_gbfi')
        )

    if not agg_tsr_df.empty:
        unified = unified.merge(
            agg_tsr_df.drop(columns=['id', 'created_at'], errors='ignore'),
            on='chainage_id',
            how='left',
            suffixes=('', '_tsr')
        )

    if not agg_tqi_df.empty:
        unified = unified.merge(
            agg_tqi_df.drop(columns=['id', 'created_at'], errors='ignore'),
            on='chainage_id',
            how='left',
            suffixes=('', '_tqi')
        )

    # Export unified dataset
    if format == 'geojson':
        unified_path = os.path.join(output_dir, "unified_rail_data.geojson")
        unified.to_file(unified_path, driver='GeoJSON')
    elif format == 'shapefile':
        unified_path = os.path.join(output_dir, "unified_rail_data.shp")
        unified.to_file(unified_path)
    elif format == 'geoparquet':
        unified_path = os.path.join(output_dir, "unified_rail_data.parquet")
        unified.to_parquet(unified_path)

    output_files['unified'] = unified_path

    # Generate a summary README
    readme_path = os.path.join(output_dir, "README_PowerBI.md")
    with open(readme_path, 'w') as f:
        f.write("# Rail Dog Power BI Export\n\n")
        f.write(f"Generated: {datetime.now().isoformat()}\n\n")
        f.write("## Files\n\n")
        f.write("### Individual Tables (CSV)\n")
        f.write("- `agg_assets.csv`: Asset counts per segment\n")
        f.write("- `agg_gbfi.csv`: Ballast fouling metrics per segment\n")
        f.write("- `agg_tsr.csv`: Temporary speed restriction data per segment\n")
        f.write("- `agg_tqi.csv`: Track quality indicator data per segment\n\n")
        f.write("### Geospatial Data\n")
        f.write(f"- `rail_segments.{format}`: Rail segments with geometry\n")
        f.write(f"- `unified_rail_data.{format}`: All metrics joined to segments\n\n")
        f.write("## Power BI Usage\n\n")
        f.write("### Option 1: Relational Model (Recommended)\n")
        f.write("1. Import `rail_segments.geojson` as the main geometry table\n")
        f.write("2. Import each CSV file as separate tables\n")
        f.write("3. Create relationships based on `chainage_id` field\n")
        f.write("4. Benefits: Smaller file sizes, easier to update individual metrics\n\n")
        f.write("### Option 2: Unified Dataset\n")
        f.write("1. Import `unified_rail_data.geojson` directly\n")
        f.write("2. Benefits: Single table, simpler for quick visualizations\n\n")
        f.write("## Visualization Ideas\n\n")
        f.write("### Track Condition Heatmap\n")
        f.write("- Color segments by TQI value (green=good, red=poor)\n")
        f.write("- Use map visual with line geometries\n")
        f.write("- Filter by line_code to focus on specific lines\n\n")
        f.write("### Ballast Fouling Analysis\n")
        f.write("- Visualize `fouled_zone` boolean on map\n")
        f.write("- Show `avg_of_avg` values as color gradient\n")
        f.write("- Identify segments requiring ballast renewal\n\n")
        f.write("### Asset Distribution\n")
        f.write("- Plot segments with `fixed_asset=true` in different color\n")
        f.write("- Size markers by asset counts (level_crossing, turnout, bridge, irj)\n\n")
        f.write("### TSR Timeline\n")
        f.write("- Time series showing TSR counts by year (cnt_2022-2026)\n")
        f.write("- Highlight segments with `open_tsr=true`\n")
        f.write("- Show `open_tsr_days` to prioritize investigations\n\n")
        f.write("### Multi-Metric Dashboard\n")
        f.write("- Combine TQI, GBFI, TSR, and asset data in single view\n")
        f.write("- Use slicers for line_code, station, chainage ranges\n")
        f.write("- Create composite risk score from multiple metrics\n\n")

    output_files['readme'] = readme_path

    logging.info(f"Power BI exports saved to: {output_dir}")
    logging.info(f"Generated {len(output_files)} files")

    return output_files


def generate_condition_summary(engine: Engine) -> pd.DataFrame:
    """
    Generate a summary of track conditions across all segments.

    Returns a DataFrame with aggregated statistics useful for executive reporting.
    """
    logging.info("Generating condition summary...")

    # Load all data
    segments_df = get_table_data(engine, table_name="rail_segments")
    agg_gbfi_df = get_table_data(engine, table_name="agg_gbfi")
    agg_tsr_df = query_agg_tsr(engine)
    agg_tqi_df = get_table_data(engine, table_name="agg_tqi")

    # Join on chainage_id
    merged = segments_df.merge(agg_gbfi_df, on='chainage_id', how='left', suffixes=('', '_gbfi'))
    merged = merged.merge(agg_tsr_df, on='chainage_id', how='left', suffixes=('', '_tsr'))
    merged = merged.merge(agg_tqi_df, on='chainage_id', how='left', suffixes=('', '_tqi'))

    # Group by line_code and generate summary stats
    summary = merged.groupby('line_code').agg({
        'chainage_id': 'count',
        'segment_length_m': 'sum',
        'fouled_zone': lambda x: x.sum() if x.notna().any() else 0,
        'open_tsr': lambda x: x.sum() if x.notna().any() else 0,
        'tqi': ['mean', 'min', 'max'],
    }).round(2)

    summary.columns = [
        'total_segments',
        'total_length_m',
        'fouled_segments',
        'segments_with_open_tsr',
        'avg_tqi',
        'min_tqi',
        'max_tqi'
    ]

    # Add percentages
    summary['pct_fouled'] = (summary['fouled_segments'] / summary['total_segments'] * 100).round(1)
    summary['pct_open_tsr'] = (summary['segments_with_open_tsr'] / summary['total_segments'] * 100).round(1)

    return summary.reset_index()
