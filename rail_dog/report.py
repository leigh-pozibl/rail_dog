# Backward compatibility shim — report has moved to rail_dog.report package.
# This file re-exports the public API so existing imports continue to work.
from rail_dog.report.report import Report
from rail_dog.report.rule_engine import RuleEngine
from rail_dog.report.excel import add_merged_header_row

__all__ = ["Report", "RuleEngine", "add_merged_header_row"]

# ── original source retained below for reference only ────────────────────────
# (not imported; kept so git history remains readable)

import logging
import os
from typing import Optional, List, Any

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import Engine

from rail_dog.utils.db_utils import get_table_data, query_agg_tsr
from rail_dog.configs.thresholds import PLAINLINE_TREATMENT_THRESHOLDS, LX_TREATMENT_THRESHOLDS, IRJ_TREATMENT_THRESHOLDS, TURNOUT_TREATMENT_THRESHOLDS, BRIDGE_TREATMENT_THRESHOLDS, PRIORITY_THRESHOLDS
from rail_dog.configs.library import REGION_PRIORITY


class _OriginalRuleEngine:
    """
    Rule engine for applying conditional logic to DataFrames using list-based rules.

    Rule format:
        rules = [
            [condition, result],
            [condition, result],
            ...
            default_result  # Optional default (can be string or None)
        ]

    Condition syntax:
        - Simple: (field, operator, value)
          Example: ("fixed_asset", "==", "Yes")

        - Qualified fields: (table.field, operator, value)
          Example: ("agg_asset.fixed_asset", "==", "Yes")
          Looks up value from export_data["agg_asset"] for the current chainage_id

        - AND: ["AND", condition1, condition2, ...]
          Example: ["AND", ("status", "==", "poor"), ("gbfi", ">=", 100)]

        - OR: ["OR", condition1, condition2, ...]
          Example: ["OR", ("status", "==", "poor"), ("status", "==", "critical")]

        - Nested: Combine AND/OR arbitrarily
          Example: ["AND",
                     ["OR", ("status", "==", "poor"), ("status", "==", "critical")],
                     ("gbfi", ">=", 100)]

    Operators:
        "==", "!=", ">", ">=", "<", "<=", "in", "not in", "is", "is not"

    Example usage:
        rules = [
            [("agg_asset.fixed_asset", "==", "Yes"), "No Treatment"],
            [["AND",
              ["OR", ("agg_tqi.status", "==", "poor"), ("agg_tqi.status", "==", "critical")],
              ("agg_gbfi.gbfi_avg", ">=", 100)
             ], "Treatment Required"],
            "No Treatment"  # Default
        ]

        engine = RuleEngine(export_data=export_data, thresholds={"min_gbfi": 100})
        result = engine.apply_rules(df, rules)
    """

    def __init__(self, export_data: Optional[dict] = None, thresholds: Optional[dict] = None):
        """
        Args:
            export_data: Optional dict of DataFrames (e.g., {"agg_asset": df, "agg_gbfi": df})
                        Used to resolve qualified field names like "agg_asset.fixed_asset"
            thresholds: Optional dict of threshold values that can be referenced in rules
        """
        self.export_data = export_data or {}
        self.thresholds = thresholds or {}

    def _get_field_value(self, row: pd.Series, field: str) -> Any:
        """
        Get field value from row, supporting qualified field names.

        OPTIMIZED: First checks if field exists directly in row (pre-merged data),
        then falls back to table lookup if needed.

        Args:
            row: Current row being evaluated
            field: Field name, either "field" or "table.field"

        Returns:
            Field value, or pd.NA if not found
        """
        # OPTIMIZATION: Check if field exists directly in row first (pre-merged data)
        if field in row.index:
            return row[field]

        # Check if field is qualified (contains a dot)
        if "." in field:
            table_name, field_name = field.split(".", 1)

            # Check if this table exists in export_data
            if table_name not in self.export_data:
                raise ValueError(f"Table '{table_name}' not found in export_data. Available: {list(self.export_data.keys())}")

            # Get chainage_id from current row
            if "chainage_id" not in row.index:
                raise ValueError(f"Row must have 'chainage_id' field to look up qualified field '{field}'")

            chainage_id = row["chainage_id"]
            source_df = self.export_data[table_name]

            # Look up the value from the source table
            if "chainage_id" not in source_df.columns:
                raise ValueError(f"Table '{table_name}' must have 'chainage_id' column")

            # Find matching row(s) in source table
            matching_rows = source_df[source_df["chainage_id"] == chainage_id]

            if matching_rows.empty:
                # No matching chainage_id in source table
                return pd.NA

            if len(matching_rows) > 1:
                logging.warning(f"Multiple rows found for chainage_id={chainage_id} in table '{table_name}', using first")

            # Get the field value
            if field_name not in matching_rows.columns:
                return pd.NA

            return matching_rows.iloc[0][field_name]

        else:
            # Simple field name - not found
            return pd.NA

    def _eval_condition(self, row: pd.Series, condition) -> bool:
        """Evaluate a single condition against a row."""
        if isinstance(condition, list):
            operator = condition[0]

            if operator == "AND":
                return all(self._eval_condition(row, c) for c in condition[1:])
            elif operator == "OR":
                return any(self._eval_condition(row, c) for c in condition[1:])
            else:
                raise ValueError(f"Unknown list operator: {operator}. Use 'AND' or 'OR'")

        elif isinstance(condition, tuple):
            if len(condition) != 3:
                raise ValueError(f"Condition tuple must have 3 elements: (field, op, value). Got: {condition}")

            field, op, value = condition

            # Handle threshold references
            if isinstance(value, str) and value.startswith("threshold."):
                threshold_key = value.split(".", 1)[1]
                if threshold_key not in self.thresholds:
                    raise ValueError(f"Threshold '{threshold_key}' not found in thresholds dict")
                value = self.thresholds[threshold_key]

            # Get field value using the qualified field resolver
            field_value = self._get_field_value(row, field)

            # Evaluate operator
            if op == "==":
                # Handle boolean comparisons - pandas/numpy bools need special handling
                if isinstance(value, bool):
                    # Convert field_value to Python bool for comparison
                    if pd.isna(field_value):
                        return False
                    # Use bool() to convert numpy.bool_ or other types to Python bool
                    return bool(field_value) == value
                return field_value == value
            elif op == "!=":
                # Handle boolean comparisons
                if isinstance(value, bool):
                    if pd.isna(field_value):
                        return True
                    return bool(field_value) != value
                return field_value != value
            elif op == ">":
                return pd.notna(field_value) and field_value > value
            elif op == ">=":
                return pd.notna(field_value) and field_value >= value
            elif op == "<":
                return pd.notna(field_value) and field_value < value
            elif op == "<=":
                return pd.notna(field_value) and field_value <= value
            elif op == "in":
                return field_value in value
            elif op == "not in":
                return field_value not in value
            elif op == "is":
                if value is None or value == "None":
                    return pd.isna(field_value)
                return field_value is value
            elif op == "is not":
                if value is None or value == "None":
                    return pd.notna(field_value)
                return field_value is not value
            else:
                raise ValueError(f"Unknown operator: {op}")

        else:
            raise ValueError(f"Condition must be a tuple or list. Got: {type(condition)}")

    def _apply_rules_to_row(self, row: pd.Series, rules: list) -> Any:
        """Apply rules to a single row and return the first matching result."""
        for rule in rules:
            # Check if this is the default (last item with no condition)
            if not isinstance(rule, (list, tuple)):
                return rule

            # Rule is [condition, result]
            if len(rule) != 2:
                raise ValueError(f"Rule must be [condition, result] or a default value. Got: {rule}")

            condition, result = rule

            if self._eval_condition(row, condition):
                return result

        # No rule matched and no default provided
        return None

    def apply_rules(self, df: pd.DataFrame, rules: list, debug: bool = False) -> pd.Series:
        """
        Apply rules to each row of a DataFrame.

        Args:
            df: DataFrame to apply rules to
            rules: List of [condition, result] pairs, optionally ending with a default value
            debug: If True, log which rules are matching (only for first row)

        Returns:
            Series with result for each row (indexed same as df)
        """
        if debug and len(df) > 0:
            logging.info("=== Rule Debug Mode (First Row) ===")
            first_row = df.iloc[0]
            logging.info(f"Evaluating chainage_id: {first_row.get('chainage_id', 'N/A')}")

            for i, rule in enumerate(rules):
                if not isinstance(rule, (list, tuple)):
                    logging.info(f"Rule {i+1}: DEFAULT -> {rule}")
                    break

                condition, result = rule
                matches = self._eval_condition(first_row, condition)
                logging.info(f"Rule {i+1}: {matches} -> {result}")
                if matches:
                    logging.info(f"  MATCHED! Returning: {result}")
                    break

        return df.apply(lambda row: self._apply_rules_to_row(row, rules), axis=1)

    def debug_row(self, row: pd.Series, rules: list) -> None:
        """
        Debug helper to show which rules match for a specific row.

        Args:
            row: Single row from DataFrame
            rules: List of rules to evaluate
        """
        print(f"\n=== Debugging Row: {row.get('chainage_id', 'N/A')} ===")

        for i, rule in enumerate(rules):
            if not isinstance(rule, (list, tuple)):
                print(f"Rule {i+1}: DEFAULT -> {rule}")
                break

            condition, result = rule
            try:
                matches = self._eval_condition(row, condition)
                print(f"Rule {i+1}: {'✓ MATCH' if matches else '✗ no match'} -> {result}")
                if matches:
                    print(f"  Result: {result}")
                    break
            except Exception as e:
                print(f"Rule {i+1}: ERROR - {e}")


class _OriginalReport():
    def __init__(
        self,
        engine: Engine,
        output_path: str,
        analysis_dates: Optional[dict] = None,
        global_date=None,
        mapbox_token: str = "",
    ):
        """
        Args:
            engine: database engine
            output_path: Path to save the Excel file
            analysis_dates: Per-type dates keyed by data type, e.g.
                            {"GBFI": datetime(...), "TQI": datetime(...), "TG": datetime(...), ...}
                            Typically sourced from base_data.analysis_dates in the config.
            global_date: Fallback date used for any type not present in analysis_dates.
                         If neither is provided no date filter is applied.
        """
        self.engine = engine
        self.output_path = output_path
        self.analysis_dates = analysis_dates or {}
        self.global_date = global_date
        self.mapbox_token = mapbox_token

        self.export_data = {}

        self.cols_list = list()  # Will be set after merging data, used for column letter lookups in formulas

        self.generate_report()

    def _date_query(self, table: str, key: str, cols: str = "*") -> str:
        """Build a SELECT query with an optional collection_date filter.

        - Specific date in analysis_dates: exact match on collection_date.
        - global_date fallback: most recent record per chainage_id up to and
          including global_date (DISTINCT ON keeps only the latest row).
        - No date configured: no filter, all rows returned.
        """
        if key in self.analysis_dates:
            date_str = str(self.analysis_dates[key])[:10]
            return f"SELECT {cols} FROM {table} WHERE collection_date = '{date_str}'"

        if self.global_date:
            date_str = str(self.global_date)[:10]
            return (
                f"SELECT {cols} FROM {table} "
                f"WHERE collection_date = ("
                f"SELECT MAX(collection_date) FROM {table} WHERE collection_date <= '{date_str}'"
                f")"
            )

        return f"SELECT {cols} FROM {table}"

    def _check_date_load(self, df: pd.DataFrame, table: str, key: str):
        """Raise if a date-filtered query returned no rows."""
        if not df.empty:
            return
        if key in self.analysis_dates:
            date_str = str(self.analysis_dates[key])[:10]
            raise ValueError(
                f"No data found in '{table}' for analysis date {date_str} (key '{key}'). "
                f"Check analysis_dates config or verify data has been loaded for this date."
            )
        if self.global_date:
            date_str = str(self.global_date)[:10]
            raise ValueError(
                f"No data found in '{table}' up to global_date {date_str}. "
                f"Verify data has been loaded for this date."
            )

    def find_plainline_treatment(self, segments_df: pd.DataFrame) -> pd.Series:
        """
        Apply plainline treatment recommendation rules to segment data.

        Uses qualified field names (table.field) to reference fields from different
        aggregated tables in self.export_data.

        Original Excel formula:
        =IF(
            M2="Yes",
            "No Treatment",
            IF(
                AND(
                    OR(AK2="poor", AK2="critical"),
                    S2>=Thresholds!$T$8
                ),
                "Local Formation Renewal (Bog Hole)",
                IF(
                    AND(
                        S2>=Thresholds!$T$9,
                        OR(AU2="s1", AU2="s2"),
                        AC2="Yes"
                    ),
                    "Local Formation Renewal (Bog Hole)",
                    IF(
                        AND(
                            OR(AK2="poor", AK2="critical"),
                            T2>=Thresholds!$V$10
                        ),
                        "Ballast Clean",
                        IF(
                            AND(
                                OR(AK2="poor", AK2="critical"),
                                T2<Thresholds!$V$11,
                                Y2<Thresholds!$Z$11
                            ),
                            "Lift & Tamp",
                            IF(
                                AND(
                                    OR(AK2="poor", AK2="critical"),
                                    T2<Thresholds!$V$12,
                                    Y2>=Thresholds!$Z$12
                                ),
                                "Tamp Only",
                                "No Treatment"
                            )
                        )
                    )
                )
            )
        )

        Column mappings:
        - M2 = fixed_asset (agg_asset.fixed_asset)
        - AK2 = TQI status (agg_tqi.status)
        - S2 = GBFI max_of_avg (agg_gbfi.max_of_avg)
        - T2 = GBFI avg_of_avg (agg_gbfi.avg_of_avg)
        - Y2 = ballast_centre (agg_ballast.ballast_centre)
        - AU2 = DTR worst_dtf (agg_dtr.worst_dtf)
        - AC2 = complete_tsr (agg_tsr.complete_tsr)

        Threshold mappings:
        - Thresholds!$T$8 = gbfi_bog_threshold_1 (40)
        - Thresholds!$T$9 = gbfi_bog_threshold_2 (40)
        - Thresholds!$V$10 = ballast_clean_threshold (20)
        - Thresholds!$V$11 = lift_tamp_gbfi_avg_threshold (20)
        - Thresholds!$Z$11 = lift_tamp_ballast_depth_threshold (0.25)
        - Thresholds!$V$12 = tamp_only_gbfi_avg_threshold (20)
        - Thresholds!$Z$12 = tamp_only_ballast_depth_threshold (0.25)

        Args:
            segments_df: DataFrame with chainage_id column (can be just segments or merged data)

        Returns:
            Series with treatment recommendation for each chainage_id
        """
        plainline_treatment_rules = [
            # Rule 1: IF fixed_asset=True, "No Treatment"
            [
                ("agg_asset.fixed_asset", "==", True),
                "No Treatment"
            ],

            # Rule 2: IF AND(OR(TQI status="poor"|"critical"), GBFI>=threshold), "Formation Renewal"
            [
                ["AND",
                    ["OR",
                        ("agg_tqi.status", "==", "poor"),
                        ("agg_tqi.status", "==", "critical")
                    ],
                    ("agg_gbfi.max_of_avg", ">=", "threshold.gbfi_bog_threshold_1")
                ],
                "Local Formation Renewal (Bog Hole)"
            ],

            # Rule 3: IF AND(GBFI>=threshold, OR(DTR severity), TSR history), "Formation Renewal"
            [
                ["AND",
                    ("agg_gbfi.max_of_avg", ">=", "threshold.gbfi_bog_threshold_2"),
                    ["OR",
                        ("agg_dtr.worst_dtf", "==", "s1"),
                        ("agg_dtr.worst_dtf", "==", "s2")
                    ],
                    ("agg_tsr.complete_tsr", "==", True)
                ],
                "Local Formation Renewal (Bog Hole)"
            ],

            # Rule 4: IF AND(OR(TQI status="poor"|"critical"), GBFI Avg), "Ballast Clean"
            [
                ["AND",
                    ["OR",
                        ("agg_tqi.status", "==", "poor"),
                        ("agg_tqi.status", "==", "critical")
                    ],
                    ("agg_gbfi.avg_of_avg", ">=", "threshold.ballast_clean_threshold")
                ],
                "Ballast Clean"
            ],

            # Rule 5: IF AND(OR(TQI status="poor"|"critical"), GBFI Avg, Ballast Centre, "Lift & Tamp"
            [
                ["AND",
                    ["OR",
                        ("agg_tqi.status", "==", "poor"),
                        ("agg_tqi.status", "==", "critical")
                    ],
                    ("agg_gbfi.avg_of_avg", "<", "threshold.lift_tamp_gbfi_avg_threshold"),
                    ("agg_ballast.ballast_centre", "<", "threshold.lift_tamp_ballast_depth_threshold")
                ],
                "Lift & Tamp"
            ],

            # Rule 6: IF AND(OR(TQI status="poor"|"critical"), GBFI Avg, Ballast Cnetre), "Tamp Only"
            [
                ["AND",
                    ["OR",
                        ("agg_tqi.status", "==", "poor"),
                        ("agg_tqi.status", "==", "critical")
                    ],
                    ("agg_gbfi.avg_of_avg", "<", "threshold.tamp_only_gbfi_avg_threshold"),
                    ("agg_ballast.ballast_centre", ">=", "threshold.tamp_only_ballast_depth_threshold")
                ],
                "Tamp Only"
            ],

            # Default
            "No Treatment"
        ]

        # Apply rules using the RuleEngine with export_data for qualified field lookups
        engine = RuleEngine(export_data=self.export_data, thresholds=PLAINLINE_TREATMENT_THRESHOLDS)
        return engine.apply_rules(segments_df, plainline_treatment_rules)
    
    def find_lx_treatment(self, segments_df: pd.DataFrame) -> pd.Series:
        """
        Apply level crossing (LX) treatment recommendation rules to segment data.

        Original Excel formula:
        =IF(
            M2<>"Yes",
            "No treatment",
            IF(
                AND(
                    N2>0,
                    AC2="Yes",
                    OR(AK2="poor", AK2="critical"),
                    OR(S2>Thresholds!$T$14, T2>Thresholds!$V$14)
                ),
                N2,
                IF(
                    AND(
                        OR(AK2="poor", AK2="critical"),
                        OR(AU2="s1", AU2="s2", AU2="s3"),
                        T2>Thresholds!$V$18,
                        Y2>=Thresholds!$Z$18,
                        I2>0
                    ),
                    Thresholds!$Q$18,
                    "No treatment"
                )
            )
        )

        Column mappings:
        - M2 = fixed_asset (agg_asset.fixed_asset)
        - N2 = wz_level_crossing count (agg_asset.wz_level_crossing)
        - I2 = level_crossing count (agg_asset.level_crossing)
        - AC2 = complete_tsr (agg_tsr.complete_tsr)
        - AK2 = TQI status (agg_tqi.status)
        - S2 = GBFI max_of_avg (agg_gbfi.max_of_avg)
        - T2 = GBFI avg_of_avg (agg_gbfi.avg_of_avg)
        - Y2 = ballast_centre (agg_ballast.ballast_centre)
        - AU2 = DTR worst_dtf (agg_dtr.worst_dtf)
        - Thresholds!$T$14 = lx_renewal_gbfi_max_threshold (40)
        - Thresholds!$V$14 = lx_renewal_gbfi_avg_threshold (30)
        - Thresholds!$V$18 = tamp_only_gbfi_avg_threshold (30)
        - Thresholds!$Z$18 = tamp_only_ballast_depth_threshold (0.25)
        - Thresholds!$Q$18 = "Tamp Only" (result string)

        Args:
            segments_df: DataFrame with chainage_id column (can be just segments or merged data)

        Returns:
            Series with treatment recommendation for each chainage_id
            Note: Returns the wz_level_crossing count (numeric) when renewal conditions are met
        """
        lx_treatment_rules = [
            # Rule 1: IF fixed_asset != True, "No Treatment"
            [
                ("agg_asset.fixed_asset", "==", False),
                "No Treatment"
            ],

            # Rule 2: IF AND(WZ Level Crossing > 0, TSR History, TQI poor/critical, (GBFI max OR avg exceeds threshold)), return wz_level_crossing count
            [
                ["AND",
                    ("agg_asset.wz_level_crossing", ">", 0),
                    ("agg_tsr.complete_tsr", "==", True),
                    ["OR",
                        ("agg_tqi.status", "==", "poor"),
                        ("agg_tqi.status", "==", "critical")
                    ],
                    ["OR",
                        ("agg_gbfi.max_of_avg", ">", "threshold.lx_renewal_gbfi_max_threshold"),
                        ("agg_gbfi.avg_of_avg", ">", "threshold.lx_renewal_gbfi_avg_threshold")
                    ]
                ],
                "__LX_COUNT__"  # Placeholder to be replaced with actual count
            ],

            # Rule 3: IF AND(TQI poor/critical, DTR S1/S2/S3, GBFI avg > threshold, ballast >= threshold, level_crossing > 0), "Tamp Only"
            [
                ["AND",
                    ["OR",
                        ("agg_tqi.status", "==", "poor"),
                        ("agg_tqi.status", "==", "critical")
                    ],
                    ["OR",
                        ("agg_dtr.worst_dtf", "==", "s1"),
                        ("agg_dtr.worst_dtf", "==", "s2"),
                        ("agg_dtr.worst_dtf", "==", "s3")
                    ],
                    ("agg_gbfi.avg_of_avg", ">", "threshold.tamp_only_gbfi_avg_threshold"),
                    ("agg_ballast.ballast_centre", ">=", "threshold.tamp_only_ballast_depth_threshold"),
                    ("agg_asset.level_crossing", ">", 0)
                ],
                "Tamp Only"
            ],

            # Default
            "No Treatment"
        ]

        # Apply rules using the RuleEngine with export_data for qualified field lookups
        engine = RuleEngine(export_data=self.export_data, thresholds=LX_TREATMENT_THRESHOLDS)
        result = engine.apply_rules(segments_df, lx_treatment_rules)

        # OPTIMIZED: Vectorized replacement of __LX_COUNT__ placeholder with actual wz_level_crossing count
        mask = result == "__LX_COUNT__"
        if mask.any() and "agg_asset.wz_level_crossing" in segments_df.columns:
            result[mask] = segments_df.loc[mask, "agg_asset.wz_level_crossing"]

        return result
    
    def find_irj_treatment(self, segments_df: pd.DataFrame) -> pd.Series:
        """
        Apply IRJ (Insulated Rail Joint) treatment recommendation rules to segment data.

        Original Excel formula:

        =IF(
            M2<>"Yes",
            "No treatment",
            IF(
                AND(
                    S2 > Thresholds!$T$15,
                    OR(AK2 = "Critical", AK2 = "Poor"),
                    OR(AU2 = "s1", AU2 = "s2"),
                    J2 > 0
                ),
                J2,
                IF(
                    AND(
                        OR(AK2 = "Critical", AK2 = "Poor"),
                        OR(AU2 = "s1", AU2 = "s2", AU2 = "s3"),
                        Y2 >= Thresholds!$Z$18,
                        J2 > 0,
                        T2>Thresholds!$V$18
                    ),
                    Thresholds!$Q$18,
                    "No treatment"
                )
            )
        )

        Column mappings:
        - M2 = fixed_asset (agg_asset.fixed_asset)
        - J2 = irj count (agg_asset.irj)
        - S2 = GBFI max_of_avg (agg_gbfi.max_of_avg)
        - T2 = GBFI avg_of_avg (agg_gbfi.avg_of_avg)
        - Y2 = ballast_centre (agg_ballast.ballast_centre)
        - AK2 = TQI status (agg_tqi.status)
        - AU2 = DTR worst_dtf (agg_dtr.worst_dtf)

        Args:
            segments_df: DataFrame with chainage_id column (can be just segments or merged data)

        Returns:
            Series with treatment recommendation for each chainage_id
            Note: Returns the IRJ count (numeric) when renewal conditions are met
        """
        irj_treatment_rules = [
            # Rule 1: IF fixed_asset != True, "No Treatment"
            [
                ("agg_asset.fixed_asset", "==", False),
                "No Treatment"
            ],

            # Rule 2: IF AND(GBFI max > threshold, TQI status poor/critical, DTR S1/S2, IRJ count > 0), return IRJ count
            [
                ["AND",
                    ("agg_gbfi.max_of_avg", ">", "threshold.irj_renewal_gbfi_max_threshold"),
                    ["OR",
                        ("agg_tqi.status", "==", "critical"),
                        ("agg_tqi.status", "==", "poor")
                    ],
                    ["OR",
                        ("agg_dtr.worst_dtf", "==", "s1"),
                        ("agg_dtr.worst_dtf", "==", "s2")
                    ],
                    ("agg_asset.irj", ">", 0)
                ],
                "__IRJ_COUNT__"  # Placeholder to be replaced with actual count
            ],

            # Rule 3: IF AND(TQI status poor/critical, DTR S1/S2/S3, ballast >= threshold, IRJ > 0, GBFI avg > threshold), "Tamp Only"
            [
                ["AND",
                    ["OR",
                        ("agg_tqi.status", "==", "critical"),
                        ("agg_tqi.status", "==", "poor")
                    ],
                    ["OR",
                        ("agg_dtr.worst_dtf", "==", "s1"),
                        ("agg_dtr.worst_dtf", "==", "s2"),
                        ("agg_dtr.worst_dtf", "==", "s3")
                    ],
                    ("agg_ballast.ballast_centre", ">=", "threshold.tamp_only_ballast_depth_threshold"),
                    ("agg_asset.irj", ">", 0),
                    ("agg_gbfi.avg_of_avg", ">", "threshold.tamp_only_gbfi_avg_threshold")
                ],
                "Tamp Only"
            ],

            # Default
            "No Treatment"
        ]

        # Apply rules using the RuleEngine with export_data for qualified field lookups
        engine = RuleEngine(export_data=self.export_data, thresholds=IRJ_TREATMENT_THRESHOLDS)
        result = engine.apply_rules(segments_df, irj_treatment_rules)

        # OPTIMIZED: Vectorized replacement of __IRJ_COUNT__ placeholder with actual IRJ count
        mask = result == "__IRJ_COUNT__"
        if mask.any() and "agg_asset.irj" in segments_df.columns:
            result[mask] = segments_df.loc[mask, "agg_asset.irj"]

        return result

    def find_turnout_treatment(self, segments_df: pd.DataFrame) -> pd.Series:
        """
        Apply turnout treatment recommendation rules to segment data.

        Original Excel formula:
        =IF(
            M2="Yes",
            IF(
                AND(
                    Y2 >= Thresholds!$Z$16,
                    OR(AU2 = "s1", AU2 = "s2", AU2 = "s3"),
                    AC2 = "Yes",
                    P2 > 0
                ),
                P2,
                IF(
                    AND(
                        OR(AK2 = "Poor", AK2 = "Critical"),
                        Y2 >= Thresholds!$Z$18,
                        OR(AU2 = "s1", AU2 = "s2", AU2 = "s3"),
                        T2 > Thresholds!$V$18
                    ),
                    Thresholds!$Q$18,
                    "No treatment"
                )
            ),
            "No treatment"
        )

        Column mappings:
        - M2 = fixed_asset (agg_asset.fixed_asset)
        - P2 = wz turnout count (agg_asset.wz_turnout)
        - Y2 = ballast_centre (agg_ballast.ballast_centre)
        - T2 = GBFI avg_of_avg (agg_gbfi.avg_of_avg)
        - AK2 = TQI status (agg_tqi.status)
        - AU2 = DTR worst_dtf (agg_dtr.worst_dtf)
        - AC2 = complete_tsr (agg_tsr.complete_tsr)

        Args:
            segments_df: DataFrame with chainage_id column (can be just segments or merged data)

        Returns:
            Series with treatment recommendation for each chainage_id
            Note: Returns the turnout count (numeric) when renewal conditions are met
        """
        turnout_treatment_rules = [
            # Rule 1: IF fixed_asset != True, "No Treatment"
            [
                ("agg_asset.fixed_asset", "==", False),
                "No Treatment"
            ],

            # Rule 2: IF AND(ballast >= threshold, DTR S1/S2/S3, TSR history, wz_turnout > 0), return turnout count
            [
                ["AND",
                    ("agg_ballast.ballast_centre", ">=", "threshold.turnout_renewal_ballast_depth_threshold"),
                    ["OR",
                        ("agg_dtr.worst_dtf", "==", "s1"),
                        ("agg_dtr.worst_dtf", "==", "s2"),
                        ("agg_dtr.worst_dtf", "==", "s3")
                    ],
                    ("agg_tsr.complete_tsr", "==", True),
                    ("agg_asset.wz_turnout", ">", 0)
                ],
                "__TURNOUT_COUNT__"  # Placeholder to be replaced with actual count
            ],

            # Rule 3: IF AND(TQI poor/critical, ballast >= threshold, DTR S1/S2/S3, GBFI avg > threshold), "Tamp Only"
            [
                ["AND",
                    ["OR",
                        ("agg_tqi.status", "==", "poor"),
                        ("agg_tqi.status", "==", "critical")
                    ],
                    ("agg_ballast.ballast_centre", ">=", "threshold.tamp_only_ballast_depth_threshold"),
                    ["OR",
                        ("agg_dtr.worst_dtf", "==", "s1"),
                        ("agg_dtr.worst_dtf", "==", "s2"),
                        ("agg_dtr.worst_dtf", "==", "s3")
                    ],
                    ("agg_gbfi.avg_of_avg", ">", "threshold.tamp_only_gbfi_avg_threshold")
                ],
                "Tamp Only"
            ],

            # Default
            "No Treatment"
        ]

        # Apply rules using the RuleEngine with export_data for qualified field lookups
        engine = RuleEngine(export_data=self.export_data, thresholds=TURNOUT_TREATMENT_THRESHOLDS)
        result = engine.apply_rules(segments_df, turnout_treatment_rules)

        # OPTIMIZED: Vectorized replacement of __TURNOUT_COUNT__ placeholder with actual turnout count
        mask = result == "__TURNOUT_COUNT__"
        if mask.any() and "agg_asset.wz_turnout" in segments_df.columns:
            result[mask] = segments_df.loc[mask, "agg_asset.wz_turnout"]

        return result

    def find_bridge_treatment(self, segments_df: pd.DataFrame) -> pd.Series:
        """
        Apply bridge treatment recommendation rules to segment data.

        Original Excel formula:
        =IF(
            M2<>"Yes",
            "No treatment",
            IF(
                AND(
                    Y2 >= Thresholds!$Z$17,
                    OR(AU2="s1", AU2="s2"),
                    T2 > Thresholds!$V$17,
                    AC2 = "Yes",
                    Q2 > 0
                ),
                Q2,
                IF(
                    AND(
                        OR(AK2="Poor", AK2="Critical"),
                        Y2 >= Thresholds!$Z$18,
                        OR(AU2="s1", AU2="s2", AU2="s3"),
                        T2 > Thresholds!$V$18
                    ),
                    Thresholds!$Q$18,
                    "No treatment"
                )
            )
        )

        Column mappings:
        - M2 = fixed_asset (agg_asset.fixed_asset)
        - Q2 = bridge count (agg_asset.wz_bridge)
        - Y2 = ballast_centre (agg_ballast.ballast_centre)
        - T2 = GBFI avg_of_avg (agg_gbfi.avg_of_avg)
        - AC2 = complete_tsr (agg_tsr.complete_tsr)
        - AU2 = DTR worst_dtf (agg_dtr.worst_dtf)
        - AK2 = TQI status (agg_tqi.status)

        Threshold mappings:
        - Thresholds!$Z$17 = bridge_renewal_ballast_depth_threshold (0.25)
        - Thresholds!$V$17 = bridge_renewal_gbfi_avg_threshold (30)
        - Thresholds!$V$18 = tamp_only_gbfi_avg_threshold (30)
        - Thresholds!$Z$18 = tamp_only_ballast_depth_threshold (0.25)
        - Thresholds!$Q$18 = "Tamp Only" (result string)

        Args:
            segments_df: DataFrame with chainage_id column (can be just segments or merged data)

        Returns:
            Series with treatment recommendation for each chainage_id
            Note: Returns the bridge count (numeric) when renewal conditions are met
        """
        bridge_treatment_rules = [
            # Rule 1: IF fixed_asset != True, "No Treatment"
            [
                ("agg_asset.fixed_asset", "==", False),
                "No Treatment"
            ],

            # Rule 2: IF AND(ballast >= threshold, DTR S1/S2, GBFI avg > threshold, TSR history, bridge > 0), return bridge count
            [
                ["AND",
                    ("agg_ballast.ballast_centre", ">=", "threshold.bridge_renewal_ballast_depth_threshold"),
                    ["OR",
                        ("agg_dtr.worst_dtf", "==", "s1"),
                        ("agg_dtr.worst_dtf", "==", "s2")
                    ],
                    ("agg_gbfi.avg_of_avg", ">", "threshold.bridge_renewal_gbfi_avg_threshold"),
                    ("agg_tsr.complete_tsr", "==", True),
                    ("agg_asset.wz_bridge", ">", 0)
                ],
                "__BRIDGE_COUNT__"  # Placeholder to be replaced with actual count
            ],

            # Rule 3: IF AND(TQI poor/critical, ballast >= threshold, DTR S1/S2/S3, GBFI avg > threshold), "Tamp Only"
            [
                ["AND",
                    ["OR",
                        ("agg_tqi.status", "==", "poor"),
                        ("agg_tqi.status", "==", "critical")
                    ],
                    ("agg_ballast.ballast_centre", ">=", "threshold.tamp_only_ballast_depth_threshold"),
                    ["OR",
                        ("agg_dtr.worst_dtf", "==", "s1"),
                        ("agg_dtr.worst_dtf", "==", "s2"),
                        ("agg_dtr.worst_dtf", "==", "s3")
                    ],
                    ("agg_gbfi.avg_of_avg", ">", "threshold.tamp_only_gbfi_avg_threshold")
                ],
                "Tamp Only"
            ],

            # Default
            "No Treatment"
        ]

        # Apply rules using the RuleEngine with export_data for qualified field lookups
        engine = RuleEngine(export_data=self.export_data, thresholds=BRIDGE_TREATMENT_THRESHOLDS)
        result = engine.apply_rules(segments_df, bridge_treatment_rules)

        # OPTIMIZED: Vectorized replacement of __BRIDGE_COUNT__ placeholder with actual bridge count
        mask = result == "__BRIDGE_COUNT__"
        if mask.any() and "agg_asset.wz_bridge" in segments_df.columns:
            result[mask] = segments_df.loc[mask, "agg_asset.wz_bridge"]

        return result

    def generate_report(
        self,
        line_codes: Optional[List[str]] = None
    ):
        """
        Retrieve all relevant data required in order to generate the final xlsx report.set
        pack data into export_data dict
        """
        # If no line codes specified, get all unique line codes from segments
        if line_codes is None:
            segments_df = get_table_data(self.engine, table_name="rail_segments")
            line_codes = sorted(segments_df['line_code'].unique())
            
        self.line_codes = line_codes
        logging.info(f"Found line codes: {self.line_codes}")

        # Load all data tables once
        logging.info("Loading data from database...")
        segments_df = get_table_data(self.engine, table_name="rail_segments")
        agg_asset_df = get_table_data(self.engine, table_name="agg_assets")
        agg_gbfi_df = get_table_data(self.engine, query=self._date_query("agg_gbfi", "GBFI"))
        self._check_date_load(agg_gbfi_df, "agg_gbfi", "GBFI")
        agg_ballast_df = get_table_data(self.engine, query=self._date_query("agg_ballast", "BALL"))
        self._check_date_load(agg_ballast_df, "agg_ballast", "BALL")
        agg_moisture_df = get_table_data(self.engine, query=self._date_query("agg_moisture", "MOI"))
        self._check_date_load(agg_moisture_df, "agg_moisture", "MOI")
        agg_tsr_df = query_agg_tsr(self.engine, as_of_date=self.analysis_dates.get("TSR") or self.global_date)
        agg_tqi_df = get_table_data(self.engine, query=self._date_query("agg_tqi", "TQI"))
        self._check_date_load(agg_tqi_df, "agg_tqi", "TQI")
        agg_tqi_all_df = get_table_data(self.engine, table_name="agg_tqi")

        # Trend fields are stored directly in agg_tqi per collection_date.
        # Ensure columns are present in case TQI data has no trend yet.
        _trend_cols = ['trend', 'trend_slope', 'trend_r_squared',
                       'step_change_type', 'step_change_date', 'trend_segment_start']
        for _col in _trend_cols:
            if _col not in agg_tqi_df.columns:
                agg_tqi_df[_col] = None
            if _col not in agg_tqi_all_df.columns:
                agg_tqi_all_df[_col] = None
        agg_dtr_df = get_table_data(self.engine, query=self._date_query("agg_dtr", "DTR"))
        self._check_date_load(agg_dtr_df, "agg_dtr", "DTR")
        agg_tg_df = get_table_data(self.engine, query=self._date_query("agg_tg", "TG", cols="chainage_id, avg_speed"))
        self._check_date_load(agg_tg_df, "agg_tg", "TG")

        # Remove geometry and metadata columns from segments for Excel export
        self.export_data["segments"] = segments_df.drop(columns=['geometry', 'created_at', 'id'], errors='ignore')
        
        # Map avg_speed from agg_tg onto segments (column always present, populated when data available)
        self.export_data["segments"]['speed'] = None
        if not agg_tg_df.empty and 'avg_speed' in agg_tg_df.columns:
            speed_map = agg_tg_df.set_index('chainage_id')['avg_speed']
            self.export_data["segments"]['speed'] = self.export_data["segments"]['chainage_id'].map(speed_map)

        # Remove id and created_at from agg tables
        self.export_data["agg_asset"] = agg_asset_df.drop(columns=['id', 'created_at'], errors='ignore')
        # Also drop ballast_centre/ballast_lt_250mm - these should only come from agg_ballast
        self.export_data["agg_gbfi"] = agg_gbfi_df.drop(columns=['id', 'collection_date', 'created_at', 'ballast_centre', 'ballast_lt_250mm'], errors='ignore')
        self.export_data["agg_ballast"] = agg_ballast_df.drop(columns=['id', 'collection_date', 'created_at'], errors='ignore')
        self.export_data["agg_moisture"] = agg_moisture_df[['chainage_id', 'max_of_avg', 'avg_of_avg']] if not agg_moisture_df.empty else agg_moisture_df
        self.export_data["agg_tsr"] = agg_tsr_df.drop(columns=['id', 'created_at'], errors='ignore')
        self.export_data["agg_tqi"] = agg_tqi_df.drop(columns=['id', 'collection_date', 'created_at'], errors='ignore')
        # Explicit column order — formula in write_tqi_trend_tab assumes A=chainage_id, B=collection_date, C=tqi
        # Join lat/lng from segments for map support
        seg_coords = segments_df[['chainage_id', 'mid_coord_lat', 'mid_coord_lng']].drop_duplicates('chainage_id')
        self.export_data["agg_tqi_all"] = (
            agg_tqi_all_df[['chainage_id', 'collection_date', 'tqi', 'status',
                            'trend', 'trend_slope', 'trend_r_squared',
                            'step_change_type', 'step_change_date', 'trend_segment_start']]
            .merge(seg_coords, on='chainage_id', how='left')
            .sort_values(['chainage_id', 'collection_date'])
            .reset_index(drop=True)
        )
        self.export_data["agg_dtr"] = agg_dtr_df.drop(columns=['id', 'collection_date', 'created_at'], errors='ignore')

        # OPTIMIZATION: Pre-merge all aggregated data with segments for faster rule evaluation
        logging.info("Merging all data tables...")
        merged_df = segments_df[['chainage_id']].copy()

        # Add asset fields
        merged_df = merged_df.merge(
            self.export_data["agg_asset"].add_prefix('agg_asset.'),
            left_on='chainage_id',
            right_on='agg_asset.chainage_id',
            how='left'
        ).drop(columns=['agg_asset.chainage_id'], errors='ignore')

        # Add GBFI fields
        merged_df = merged_df.merge(
            self.export_data["agg_gbfi"].add_prefix('agg_gbfi.'),
            left_on='chainage_id',
            right_on='agg_gbfi.chainage_id',
            how='left'
        ).drop(columns=['agg_gbfi.chainage_id'], errors='ignore')
        
        # Add ballast fields
        merged_df = merged_df.merge(
            self.export_data["agg_ballast"].add_prefix('agg_ballast.'),
            left_on='chainage_id',
            right_on='agg_ballast.chainage_id',
            how='left'
        ).drop(columns=['agg_ballast.chainage_id'], errors='ignore')

        # Add moisture fields
        if not self.export_data["agg_moisture"].empty:
            merged_df = merged_df.merge(
                self.export_data["agg_moisture"].add_prefix('agg_moisture.'),
                left_on='chainage_id',
                right_on='agg_moisture.chainage_id',
                how='left'
            ).drop(columns=['agg_moisture.chainage_id'], errors='ignore')

        # Add TSR fields
        merged_df = merged_df.merge(
            self.export_data["agg_tsr"].add_prefix('agg_tsr.'),
            left_on='chainage_id',
            right_on='agg_tsr.chainage_id',
            how='left'
        ).drop(columns=['agg_tsr.chainage_id'], errors='ignore')

        # Add TQI fields
        merged_df = merged_df.merge(
            self.export_data["agg_tqi"].add_prefix('agg_tqi.'),
            left_on='chainage_id',
            right_on='agg_tqi.chainage_id',
            how='left'
        ).drop(columns=['agg_tqi.chainage_id'], errors='ignore')

        # Add DTR fields
        merged_df = merged_df.merge(
            self.export_data["agg_dtr"].add_prefix('agg_dtr.'),
            left_on='chainage_id',
            right_on='agg_dtr.chainage_id',
            how='left'
        ).drop(columns=['agg_dtr.chainage_id'], errors='ignore')

        # Generate model recommendations using pre-merged data
        logging.info("Generating recommendations...")
        recommendations_df = pd.DataFrame({
            'chainage_id': segments_df['chainage_id'],
            'plainline_treatment_model': self.find_plainline_treatment(merged_df),
            'plainline_treatment_formula': None,  # TEMP: Excel formula for troubleshooting
            'plainline_treatment_override': None,  # Empty column for manual input
            'plainline_treatment_final': None,  # Will be replaced with Excel formula
            'lx_treatment_model': self.find_lx_treatment(merged_df),
            'lx_treatment_formula': None,  # TEMP: Excel formula for troubleshooting
            'lx_treatment_override': None,  # Empty column for manual input
            'lx_treatment_final': None,  # Will be replaced with Excel formula
            'irj_treatment_model': self.find_irj_treatment(merged_df),
            'irj_treatment_formula': None,  # TEMP: Excel formula for troubleshooting
            'irj_treatment_override': None,  # Empty column for manual input
            'irj_treatment_final': None,  # Will be replaced with Excel formula
            'turnout_treatment_model': self.find_turnout_treatment(merged_df),
            'turnout_treatment_formula': None,  # TEMP: Excel formula for troubleshooting
            'turnout_treatment_override': None,  # Empty column for manual input
            'turnout_treatment_final': None,  # Will be replaced with Excel formula
            'bridge_treatment_model': self.find_bridge_treatment(merged_df),
            'bridge_treatment_formula': None,  # TEMP: Excel formula for troubleshooting
            'bridge_treatment_override': None,  # Empty column for manual input
            'bridge_treatment_final': None,  # Will be replaced with Excel formula
        })

        recommendations_df['problem_zone'] = None  # Will be replaced with Excel formula
        recommendations_df['problem_zone_group'] = None  # Will be replaced with Excel formula
        recommendations_df['problem_zone_length_m'] = None  # Will be replaced with Excel formula

        # Priority scoring columns - will be replaced with Excel formulas
        recommendations_df['priority_tsr'] = None
        recommendations_df['priority_dtf'] = None
        recommendations_df['priority_tqi_status'] = None
        recommendations_df['priority_tqi_trend'] = None
        recommendations_df['priority_speed'] = None
        recommendations_df['priority_curve'] = None
        recommendations_df['priority_score'] = None
        recommendations_df['priority_dtf_getting_worse'] = None
        recommendations_df['priority_fixed_asset'] = None
        recommendations_df['priority_lx_score'] = None
        recommendations_df['priority_irj_score'] = None
        recommendations_df['priority_turnout_score'] = None
        recommendations_df['priority_bridge_score'] = None
        recommendations_df['priority_fixed_asset_max'] = None

        self.export_data["recommendations"] = recommendations_df

    def calculate_problem_zone_group(self, recommendations_df: pd.DataFrame) -> pd.Series:
        """
        Calculate problem_zone_group based on consecutive problem zones.

        Excel formula: =IF(BM2=0, "", IF(COUNTA($BN$1:BN1)=0, 1, IF(BM1=1, BN1, MAXIFS($BN$1:BN1, $BM$1:BM1, 1)+1)))

        Logic:
        - If problem_zone=0 → None (empty)
        - If first problem zone → group 1
        - If previous row was also problem_zone=1 → same group number
        - If previous row was problem_zone=0 → new group number (increment)

        Args:
            recommendations_df: DataFrame with problem_zone column

        Returns:
            Series with group numbers for consecutive problem zones
            
        Example:
        Row | problem_zone | problem_zone_group | Description
        ----|--------------|-------------------|-------------
            1 |      0       |       None        | No problem
            2 |      1       |        1          | Start group 1
            3 |      1       |        1          | Continue group 1
            4 |      1       |        1          | Continue group 1
            5 |      0       |       None        | No problem
            6 |      0       |       None        | No problem
            7 |      1       |        2          | Start group 2 (gap after group 1)
            8 |      1       |        2          | Continue group 2
            9 |      0       |       None        | No problem
           10 |      1       |        3          | Start group 3 (gap after group 2)

        """
        groups = []
        current_group = 0

        for idx in range(len(recommendations_df)):
            problem_zone = recommendations_df.iloc[idx]['problem_zone']

            if problem_zone == 0:
                # No problem zone, no group
                groups.append(None)
            else:
                # problem_zone == 1
                if idx == 0:
                    # First row with problem_zone=1
                    current_group = 1
                    groups.append(current_group)
                else:
                    prev_problem_zone = recommendations_df.iloc[idx-1]['problem_zone']
                    if prev_problem_zone == 1:
                        # Continue previous group
                        groups.append(current_group)
                    else:
                        # Start new group (previous was 0, now is 1)
                        current_group += 1
                        groups.append(current_group)

        return pd.Series(groups, index=recommendations_df.index)

    def calculate_problem_zone_length(
        self,
        recommendations_df: pd.DataFrame,
        segments_df: pd.DataFrame
    ) -> pd.Series:
        """
        Calculate total chainage length (in meters) for each problem_zone_group.
        Each row in a group gets the same total length value.

        Args:
            recommendations_df: DataFrame with problem_zone_group column
            segments_df: DataFrame with chainage_start_km and chainage_end_km

        Returns:
            Series with group total lengths in meters
            
        Example:
        Row | chainage_id | problem_zone | problem_zone_group | segment_length | problem_zone_length_m
        ----|-------------|--------------|-------------------|----------------|----------------------
            1 | CHAIN-001   |      0       |       None        |     100m       |         None
            2 | CHAIN-002   |      1       |        1          |     100m       |         300m  (total for group 1)
            3 | CHAIN-003   |      1       |        1          |     100m       |         300m  (total for group 1)
            4 | CHAIN-004   |      1       |        1          |     100m       |         300m  (total for group 1)
            5 | CHAIN-005   |      0       |       None        |     100m       |         None
            6 | CHAIN-006   |      1       |        2          |     100m       |         200m  (total for group 2)
            7 | CHAIN-007   |      1       |        2          |     100m       |         200m  (total for group 2)
        """
        # Merge to get chainage info
        merged = recommendations_df.merge(
            segments_df[['chainage_id', 'chainage_start_km', 'chainage_end_km']],
            on='chainage_id',
            how='left'
        )

        # Calculate individual segment length in meters
        merged['segment_length_m'] = (merged['chainage_end_km'] - merged['chainage_start_km']) * 1000

        # Calculate group totals
        group_lengths = merged.groupby('problem_zone_group')['segment_length_m'].sum()

        # Map group totals back to each row
        length_series = merged['problem_zone_group'].map(group_lengths)

        # Set None for rows without a problem_zone_group
        length_series = length_series.where(merged['problem_zone_group'].notna(), other=float('nan'))

        return length_series

    def _get_actual_collection_date(self, table: str, key: str, date_col: str = "collection_date") -> str:
        """Return the actual collection date used for a given table and key.

        - Specific date in analysis_dates: returns that date directly (exact match query).
        - global_date fallback: queries MAX(date_col) <= global_date (single snapshot date).
        - No date configured: returns 'N/A (all records)'.
        """
        from sqlalchemy import text as sa_text
        if key in self.analysis_dates:
            return str(self.analysis_dates[key])[:10]
        if self.global_date:
            date_str = str(self.global_date)[:10]
            with self.engine.connect() as conn:
                result = conn.execute(
                    sa_text(f"SELECT MAX({date_col}) FROM {table} WHERE {date_col} <= '{date_str}'")
                ).scalar()
            return str(result)[:10] if result else "N/A"
        return "N/A (all records)"

    def write_data_sources_tab(self, writer: pd.ExcelWriter):
        """Write a summary tab listing each data source, its collection date, and age."""
        analysis_date_str = str(self.global_date)[:10] if self.global_date else "N/A"

        def _age(collection_date_str: str, key: str) -> str:
            if collection_date_str in ("N/A", "N/A (all records)"):
                return "N/A"
            ref = self.analysis_dates.get(key) or self.global_date
            if not ref:
                return "N/A"
            try:
                col_date = pd.to_datetime(collection_date_str).date()
                ref_date = pd.to_datetime(str(ref)[:10]).date()
                return f"{(ref_date - col_date).days} days"
            except Exception:
                return "N/A"

        sources = [
            ("rail_segments",  "rail_segments", None,  None),
            ("agg_assets",     "agg_assets",    None,  None),
            ("GBFI",           "agg_gbfi",      "GBFI", "collection_date"),
            ("Ballast",        "agg_ballast",   "BALL", "collection_date"),
            ("Moisture",       "agg_moisture",  "MOI",  "collection_date"),
            ("TSR",            "tsr_records",   "TSR",  "report_date", "<="),
            ("TQI",            "agg_tqi",       "TQI",  "collection_date"),
            ("DTR",            "agg_dtr",       "DTR",  "collection_date"),
            ("Track Geometry", "agg_tg",        "TG",   "collection_date"),
        ]

        def _record_count(table: str, key: str, date_col: str, col_date: str, count_op: str = "=") -> str:
            from sqlalchemy import text as sa_text
            try:
                if col_date in ("N/A", "N/A (all records)"):
                    with self.engine.connect() as conn:
                        return str(conn.execute(sa_text(f"SELECT COUNT(*) FROM {table}")).scalar())
                if count_op == "<=" and key:
                    configured = self.analysis_dates.get(key) or self.global_date
                    count_date = pd.to_datetime(str(configured)[:10]).strftime("%Y-%m-%d") if configured else col_date
                else:
                    count_date = col_date
                with self.engine.connect() as conn:
                    return str(conn.execute(
                        sa_text(f"SELECT COUNT(*) FROM {table} WHERE {date_col} {count_op} '{count_date}'")
                    ).scalar())
            except Exception:
                return "N/A"

        rows = []
        for label, table, key, date_col, *opts in sources:
            count_op = opts[0] if opts else "="
            if key is None:
                col_date = "N/A (all records)"
            else:
                col_date = self._get_actual_collection_date(table, key, date_col=date_col)
            rows.append({
                "Data Source":     label,
                "Table":           table,
                "Collection Date": col_date,
                "Record Count":    _record_count(table, key, date_col, col_date, count_op),
                "Age":             _age(col_date, key) if key else "N/A",
            })

        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name="Data Sources", index=False, startrow=2)

        worksheet = writer.sheets["Data Sources"]

        # Header row
        worksheet["A1"] = f"Analysis Date: {analysis_date_str}"
        worksheet["A1"].font = Font(bold=True, size=12)

        # Column widths and bold headers
        for col, width in [("A", 20), ("B", 20), ("C", 20), ("D", 15), ("E", 15)]:
            worksheet.column_dimensions[col].width = width
            worksheet[f"{col}3"].font = Font(bold=True, size=11)

    def write_thresholds_tab(self, writer: pd.ExcelWriter):
        """
        Write all treatment thresholds to a 'Thresholds' tab in the Excel report.

        Args:
            writer: ExcelWriter object
        """
        # Compile all thresholds into a single DataFrame
        thresholds_data = []

        # Plainline thresholds
        for key, value in PLAINLINE_TREATMENT_THRESHOLDS.items():
            thresholds_data.append({
                'Treatment Type': 'Plainline',
                'Threshold Name': key,
                'Value': value,
                'Description': self._get_threshold_description(key)
            })

        # LX thresholds
        for key, value in LX_TREATMENT_THRESHOLDS.items():
            thresholds_data.append({
                'Treatment Type': 'Level Crossing',
                'Threshold Name': key,
                'Value': value,
                'Description': self._get_threshold_description(key)
            })

        # IRJ thresholds
        for key, value in IRJ_TREATMENT_THRESHOLDS.items():
            thresholds_data.append({
                'Treatment Type': 'IRJ',
                'Threshold Name': key,
                'Value': value,
                'Description': self._get_threshold_description(key)
            })

        # Turnout thresholds
        for key, value in TURNOUT_TREATMENT_THRESHOLDS.items():
            thresholds_data.append({
                'Treatment Type': 'Turnout',
                'Threshold Name': key,
                'Value': value,
                'Description': self._get_threshold_description(key)
            })

        # Bridge thresholds
        for key, value in BRIDGE_TREATMENT_THRESHOLDS.items():
            thresholds_data.append({
                'Treatment Type': 'Bridge',
                'Threshold Name': key,
                'Value': value,
                'Description': self._get_threshold_description(key)
            })

        # Priority thresholds
        for key, value in PRIORITY_THRESHOLDS.items():
            thresholds_data.append({
                'Treatment Type': 'Priority',
                'Threshold Name': key,
                'Value': value,
                'Description': self._get_threshold_description(key)
            })

        # Create DataFrame and write to Excel
        thresholds_df = pd.DataFrame(thresholds_data)
        thresholds_df.to_excel(writer, sheet_name='Thresholds', index=False)

        # Format the Thresholds sheet
        worksheet = writer.sheets['Thresholds']

        # Set column widths
        worksheet.column_dimensions['A'].width = 20  # Treatment Type
        worksheet.column_dimensions['B'].width = 40  # Threshold Name
        worksheet.column_dimensions['C'].width = 15  # Value
        worksheet.column_dimensions['D'].width = 60  # Description

        # Style header row
        for col in ['A', 'B', 'C', 'D']:
            cell = worksheet[f'{col}1']
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        logging.info("Written Thresholds tab")

    def _get_col_letter(self, col_name: str) -> Optional[str]:
        # Helper to get column letter from column name
        if col_name not in self.cols_list:
            return None
        return get_column_letter(self.cols_list.index(col_name) + 1)
        
    def _get_threshold_description(self, threshold_name: str) -> str:
        """Get human-readable description for threshold names."""
        descriptions = {
            'gbfi_bog_threshold_1': 'GBFI max threshold for bog hole (1st condition)',
            'gbfi_bog_threshold_2': 'GBFI max threshold for bog hole (2nd condition)',
            'ballast_clean_threshold': 'GBFI avg threshold for ballast clean',
            'lift_tamp_gbfi_avg_threshold': 'GBFI avg threshold for lift & tamp',
            'lift_tamp_ballast_depth_threshold': 'Ballast depth threshold for lift & tamp (m)',
            'tamp_only_gbfi_avg_threshold': 'GBFI avg threshold for tamp only',
            'tamp_only_ballast_depth_threshold': 'Ballast depth threshold for tamp only (m)',
            'lx_renewal_gbfi_max_threshold': 'GBFI max threshold for level crossing renewal',
            'lx_renewal_gbfi_avg_threshold': 'GBFI avg threshold for level crossing renewal',
            'irj_renewal_gbfi_max_threshold': 'GBFI max threshold for IRJ renewal',
            'turnout_renewal_ballast_depth_threshold': 'Ballast depth threshold for turnout renewal (m)',
            'bridge_renewal_ballast_depth_threshold': 'Ballast depth threshold for bridge renewal (m)',
            'bridge_renewal_gbfi_avg_threshold': 'GBFI avg threshold for bridge renewal',
            'priority_weight_tsr': 'Priority weight for TSR score',
            'priority_weight_dtf': 'Priority weight for DTF score',
            'priority_weight_tqi_status': 'Priority weight for TQI status score',
            'priority_weight_tqi_trend': 'Priority weight for TQI trend score',
            'priority_weight_speed': 'Priority weight for max speed score',
            'priority_weight_curve': 'Priority weight for curve type score',
            'priority_norm_tsr': 'Priority normalization for TSR (max score)',
            'priority_norm_dtf': 'Priority normalization for DTF (max score)',
            'priority_norm_tqi_status': 'Priority normalization for TQI status (max score)',
            'priority_norm_tqi_trend': 'Priority normalization for TQI trend (max score)',
            'priority_norm_speed': 'Priority normalization for max speed (max score)',
            'priority_norm_curve': 'Priority normalization for curve type (max score)',
            'priority_speed_low': 'Speed threshold low (below = 0)',
            'priority_speed_high': 'Speed threshold high (above = 2)',
            'priority_speed_min': 'Speed threshold min for mid-range check',
            'priority_speed_mid': 'Speed threshold mid for mid-range check',
            'priority_lx_weight_fixed_asset': 'LX priority weight for fixed asset score',
            'priority_lx_weight_tsr': 'LX priority weight for TSR score',
            'priority_lx_weight_dtf': 'LX priority weight for DTF score',
            'priority_lx_weight_dtf_worse': 'LX priority weight for DTF getting worse score',
            'priority_lx_weight_speed': 'LX priority weight for max speed score',
            'priority_lx_norm_fixed_asset': 'LX priority normalization for fixed asset (max score)',
            'priority_lx_norm_tsr': 'LX priority normalization for TSR (max score)',
            'priority_lx_norm_dtf': 'LX priority normalization for DTF (max score)',
            'priority_lx_norm_dtf_worse': 'LX priority normalization for DTF getting worse (max score)',
            'priority_lx_norm_speed': 'LX priority normalization for max speed (max score)',
        }
        return descriptions.get(threshold_name, '')

    def _add_treatment_formula_columns(self, worksheet, merged: pd.DataFrame):
        """
        TEMP: Add Excel formulas to *_treatment_formula columns that reference Thresholds tab.
        These formulas should produce the same results as the model columns for troubleshooting.

        Args:
            worksheet: openpyxl worksheet object
            merged: DataFrame with all merged data
        """

        # Helper to create threshold reference formula
        def th(threshold_name):
            """Create INDEX/MATCH formula to look up threshold from Thresholds tab"""
            return f'INDEX(Thresholds!$C:$C,MATCH("{threshold_name}",Thresholds!$B:$B,0))'

        # Get column letters for data fields
        fixed_asset = self._get_col_letter('fixed_asset')
        status = self._get_col_letter('status')  # TQI status
        max_of_avg = self._get_col_letter('max_of_avg')  # GBFI max_of_avg
        avg_of_avg = self._get_col_letter('avg_of_avg')  # GBFI avg_of_avg
        ballast_centre = self._get_col_letter('ballast_centre')
        worst_dtf = self._get_col_letter('worst_dtf')  # DTR worst_dtf
        complete_tsr = self._get_col_letter('complete_tsr')
        wz_level_crossing = self._get_col_letter('wz_level_crossing')
        level_crossing = self._get_col_letter('level_crossing')
        irj = self._get_col_letter('irj')
        wz_turnout = self._get_col_letter('wz_turnout')
        wz_bridge = self._get_col_letter('wz_bridge')

        # Get formula column letters
        plainline_formula_col = self._get_col_letter('plainline_treatment_formula')
        lx_formula_col = self._get_col_letter('lx_treatment_formula')
        irj_formula_col = self._get_col_letter('irj_treatment_formula')
        turnout_formula_col = self._get_col_letter('turnout_treatment_formula')
        bridge_formula_col = self._get_col_letter('bridge_treatment_formula')

        # Log any missing source columns so it's clear why formulas may be skipped
        missing = [name for name, val in {
            'fixed_asset': fixed_asset, 'status': status, 'max_of_avg': max_of_avg,
            'avg_of_avg': avg_of_avg, 'ballast_centre': ballast_centre, 'worst_dtf': worst_dtf,
            'complete_tsr': complete_tsr, 'wz_level_crossing': wz_level_crossing,
            'level_crossing': level_crossing, 'irj': irj, 'wz_turnout': wz_turnout,
            'wz_bridge': wz_bridge,
        }.items() if val is None]
        if missing:
            logging.warning(f"Treatment formula columns missing source data columns: {missing}")

        # Write formulas for each row (starting at row 3: row 1 = merged header, row 2 = column names)
        for row_num in range(3, len(merged) + 3):
            # Plainline treatment formula
            if plainline_formula_col and fixed_asset and status and max_of_avg and avg_of_avg and ballast_centre and worst_dtf and complete_tsr:
                plainline_formula = (
                    f'=IF({fixed_asset}{row_num}=TRUE,"No Treatment",'
                    f'IF(AND(OR({status}{row_num}="poor",{status}{row_num}="critical"),{max_of_avg}{row_num}>={th("gbfi_bog_threshold_1")}),"Local Formation Renewal (Bog Hole)",'
                    f'IF(AND({max_of_avg}{row_num}>={th("gbfi_bog_threshold_2")},OR({worst_dtf}{row_num}="s1",{worst_dtf}{row_num}="s2"),{complete_tsr}{row_num}=TRUE),"Local Formation Renewal (Bog Hole)",'
                    f'IF(AND(OR({status}{row_num}="poor",{status}{row_num}="critical"),{avg_of_avg}{row_num}>={th("ballast_clean_threshold")}),"Ballast Clean",'
                    f'IF(AND(OR({status}{row_num}="poor",{status}{row_num}="critical"),{avg_of_avg}{row_num}<{th("lift_tamp_gbfi_avg_threshold")},{ballast_centre}{row_num}<{th("lift_tamp_ballast_depth_threshold")}),"Lift & Tamp",'
                    f'IF(AND(OR({status}{row_num}="poor",{status}{row_num}="critical"),{avg_of_avg}{row_num}<{th("tamp_only_gbfi_avg_threshold")},{ballast_centre}{row_num}>={th("tamp_only_ballast_depth_threshold")}),"Tamp Only",'
                    f'"No Treatment"))))))'
                )
                worksheet[f'{plainline_formula_col}{row_num}'] = plainline_formula

            # LX treatment formula
            if lx_formula_col and fixed_asset and wz_level_crossing and complete_tsr and status and max_of_avg and avg_of_avg and worst_dtf and ballast_centre and level_crossing:
                lx_formula = (
                    f'=IF({fixed_asset}{row_num}=FALSE,"No Treatment",'
                    f'IF(AND({wz_level_crossing}{row_num}>0,{complete_tsr}{row_num}=TRUE,OR({status}{row_num}="poor",{status}{row_num}="critical"),OR({max_of_avg}{row_num}>{th("lx_renewal_gbfi_max_threshold")},{avg_of_avg}{row_num}>{th("lx_renewal_gbfi_avg_threshold")})),{wz_level_crossing}{row_num},'
                    f'IF(AND(OR({status}{row_num}="poor",{status}{row_num}="critical"),OR({worst_dtf}{row_num}="s1",{worst_dtf}{row_num}="s2",{worst_dtf}{row_num}="s3"),{avg_of_avg}{row_num}>{th("tamp_only_gbfi_avg_threshold")},{ballast_centre}{row_num}>={th("tamp_only_ballast_depth_threshold")},{level_crossing}{row_num}>0),"Tamp Only",'
                    f'"No Treatment")))'
                )
                worksheet[f'{lx_formula_col}{row_num}'] = lx_formula

            # IRJ treatment formula
            if irj_formula_col and fixed_asset and max_of_avg and status and worst_dtf and irj and ballast_centre and avg_of_avg:
                irj_formula = (
                    f'=IF({fixed_asset}{row_num}=FALSE,"No Treatment",'
                    f'IF(AND({max_of_avg}{row_num}>{th("irj_renewal_gbfi_max_threshold")},OR({status}{row_num}="critical",{status}{row_num}="poor"),OR({worst_dtf}{row_num}="s1",{worst_dtf}{row_num}="s2"),{irj}{row_num}>0),{irj}{row_num},'
                    f'IF(AND(OR({status}{row_num}="critical",{status}{row_num}="poor"),OR({worst_dtf}{row_num}="s1",{worst_dtf}{row_num}="s2",{worst_dtf}{row_num}="s3"),{ballast_centre}{row_num}>={th("tamp_only_ballast_depth_threshold")},{irj}{row_num}>0,{avg_of_avg}{row_num}>{th("tamp_only_gbfi_avg_threshold")}),"Tamp Only",'
                    f'"No Treatment")))'
                )
                worksheet[f'{irj_formula_col}{row_num}'] = irj_formula

            # Turnout treatment formula
            if turnout_formula_col and fixed_asset and ballast_centre and worst_dtf and complete_tsr and wz_turnout and status and avg_of_avg:
                turnout_formula = (
                    f'=IF({fixed_asset}{row_num}=FALSE,"No Treatment",'
                    f'IF(AND({ballast_centre}{row_num}>={th("turnout_renewal_ballast_depth_threshold")},OR({worst_dtf}{row_num}="s1",{worst_dtf}{row_num}="s2",{worst_dtf}{row_num}="s3"),{complete_tsr}{row_num}=TRUE,{wz_turnout}{row_num}>0),{wz_turnout}{row_num},'
                    f'IF(AND(OR({status}{row_num}="poor",{status}{row_num}="critical"),{ballast_centre}{row_num}>={th("tamp_only_ballast_depth_threshold")},OR({worst_dtf}{row_num}="s1",{worst_dtf}{row_num}="s2",{worst_dtf}{row_num}="s3"),{avg_of_avg}{row_num}>{th("tamp_only_gbfi_avg_threshold")}),"Tamp Only",'
                    f'"No Treatment")))'
                )
                worksheet[f'{turnout_formula_col}{row_num}'] = turnout_formula

            # Bridge treatment formula
            if bridge_formula_col and fixed_asset and ballast_centre and worst_dtf and avg_of_avg and complete_tsr and wz_bridge and status:
                bridge_formula = (
                    f'=IF({fixed_asset}{row_num}=FALSE,"No Treatment",'
                    f'IF(AND({ballast_centre}{row_num}>={th("bridge_renewal_ballast_depth_threshold")},OR({worst_dtf}{row_num}="s1",{worst_dtf}{row_num}="s2"),{avg_of_avg}{row_num}>{th("bridge_renewal_gbfi_avg_threshold")},{complete_tsr}{row_num}=TRUE,{wz_bridge}{row_num}>0),{wz_bridge}{row_num},'
                    f'IF(AND(OR({status}{row_num}="poor",{status}{row_num}="critical"),{ballast_centre}{row_num}>={th("tamp_only_ballast_depth_threshold")},OR({worst_dtf}{row_num}="s1",{worst_dtf}{row_num}="s2",{worst_dtf}{row_num}="s3"),{avg_of_avg}{row_num}>{th("tamp_only_gbfi_avg_threshold")}),"Tamp Only",'
                    f'"No Treatment")))'
                )
                worksheet[f'{bridge_formula_col}{row_num}'] = bridge_formula

        logging.info("Added treatment formula columns referencing Thresholds tab")

    def _add_priority_columns(self, worksheet, merged: pd.DataFrame):
        """
        TEMP: Add Excel formulas to calculate priority based on treatment recommendations.

        Adds 6 numeric encoding columns (priority_tsr, priority_dtf, priority_tqi_status,
        priority_tqi_trend, priority_speed, priority_curve) and 1 aggregated weighted score column.

        Args:
            worksheet: openpyxl worksheet object
            merged: DataFrame with all merged data
            cols_list: List of column names in order
        """
        # Helper to create threshold reference formula
        def th(threshold_name):
            return f'INDEX(Thresholds!$C:$C,MATCH("{threshold_name}",Thresholds!$B:$B,0))'

        # Get column letters for source data fields
        open_tsr = self._get_col_letter('open_tsr')
        worst_dtf = self._get_col_letter('worst_dtf')
        tqi_status = self._get_col_letter('status')
        tqi_trend = self._get_col_letter('trend')
        max_speed = self._get_col_letter('speed')
        curve_type = self._get_col_letter('curve_type')
        plainline_final = self._get_col_letter('plainline_treatment_final')

        # Get column letters for priority output columns
        p_tsr = self._get_col_letter('priority_tsr')
        p_dtf = self._get_col_letter('priority_dtf')
        p_tqi_status = self._get_col_letter('priority_tqi_status')
        p_tqi_trend = self._get_col_letter('priority_tqi_trend')
        p_speed = self._get_col_letter('priority_speed')
        p_curve = self._get_col_letter('priority_curve')
        p_score = self._get_col_letter('priority_score')

        if not all([open_tsr, worst_dtf, tqi_status, tqi_trend, max_speed, curve_type,
                    plainline_final, p_tsr, p_dtf, p_tqi_status, p_tqi_trend, p_speed, p_curve, p_score]):
            logging.warning("Missing columns for priority formulas, skipping")
            return

        for row_num in range(3, len(merged) + 3):
            r = row_num

            # priority_tsr: No=0, else 1
            worksheet[f'{p_tsr}{r}'] = (
                f'=IF({open_tsr}{r}="No",0,1)'
            )

            # priority_dtf: N/A or S3=0, S2=1, S1=2
            worksheet[f'{p_dtf}{r}'] = (
                f'=_xlfn.IFS(OR({worst_dtf}{r}="",{worst_dtf}{r}="S3"),0,'
                f'{worst_dtf}{r}="S2",1,'
                f'{worst_dtf}{r}="S1",2)'
            )

            # priority_tqi_status: good/satisfactory=0, poor=1, critical=2
            worksheet[f'{p_tqi_status}{r}'] = (
                f'=_xlfn.IFS(OR({tqi_status}{r}="good",{tqi_status}{r}="satisfactory", {tqi_status}{r}=""),0,'
                f'{tqi_status}{r}="poor",1,'
                f'{tqi_status}{r}="critical",2)'
            )

            # priority_tqi_trend: blank/recent tamp=0, stable=1, degrading=2, repetition=3
            worksheet[f'{p_tqi_trend}{r}'] = (
                f'=_xlfn.IFS(OR({tqi_trend}{r}="",{tqi_trend}{r}="recent tamp", {tqi_trend}{r}="improving"),0,'
                f'{tqi_trend}{r}="stable",1,'
                f'{tqi_trend}{r}="degrading",2,'
                f'{tqi_trend}{r}="repetition",3)'
            )

            # priority_speed: threshold-based scoring
            worksheet[f'{p_speed}{r}'] = (
                f'=_xlfn.IFS({max_speed}{r}="",0,'
                f'{max_speed}{r}<={th("priority_speed_low")},0,'
                f'{max_speed}{r}<={th("priority_speed_med")},1,'
                f'{max_speed}{r}>{th("priority_speed_med")},2)'
            )

            # priority_curve: tangent=0, mild curve=1, sharp curve=2
            worksheet[f'{p_curve}{r}'] = (
                f'=_xlfn.IFS({curve_type}{r}="tangent",0,'
                f'{curve_type}{r}="mild_curve",1,'
                f'OR({curve_type}{r}="sharp_curve", {curve_type}{r}="spiral_entry", {curve_type}{r}="spiral_exit"),2)'
            )

            # Aggregated priority score: weighted normalized sum, 0 if no treatment
            worksheet[f'{p_score}{r}'] = (
                f'=IF({plainline_final}{r}<>"No Treatment",'
                f'({p_tsr}{r}/{th("priority_norm_tsr")})*{th("priority_weight_tsr")}+'
                f'({p_dtf}{r}/{th("priority_norm_dtf")})*{th("priority_weight_dtf")}+'
                f'({p_tqi_status}{r}/{th("priority_norm_tqi_status")})*{th("priority_weight_tqi_status")}+'
                f'({p_tqi_trend}{r}/{th("priority_norm_tqi_trend")})*{th("priority_weight_tqi_trend")}+'
                f'({p_speed}{r}/{th("priority_norm_speed")})*{th("priority_weight_speed")}+'
                f'({p_curve}{r}/{th("priority_norm_curve")})*{th("priority_weight_curve")},'
                f'0)'
            )

        # Get column letters for additional source data fields
        getting_worse = self._get_col_letter('getting_worse')
        fixed_asset = self._get_col_letter('fixed_asset')
        lx_final = self._get_col_letter('lx_treatment_final')
        irj_final = self._get_col_letter('irj_treatment_final')
        turnout_final = self._get_col_letter('turnout_treatment_final')
        bridge_final = self._get_col_letter('bridge_treatment_final')

        # Get column letters for additional priority output columns
        p_dtf_worse = self._get_col_letter('priority_dtf_getting_worse')
        p_fixed_asset = self._get_col_letter('priority_fixed_asset')
        p_lx_score = self._get_col_letter('priority_lx_score')
        p_irj_score = self._get_col_letter('priority_irj_score')
        p_turnout_score = self._get_col_letter('priority_turnout_score')
        p_bridge_score = self._get_col_letter('priority_bridge_score')
        p_fixed_asset_max = self._get_col_letter('priority_fixed_asset_max')

        if not all([getting_worse, fixed_asset, lx_final, irj_final, turnout_final, bridge_final,
                    p_dtf_worse, p_fixed_asset, p_lx_score, p_irj_score, p_turnout_score, p_bridge_score,
                    p_fixed_asset_max]):
            logging.warning("Missing columns for fixed asset priority formulas, skipping")
            logging.info("Added priority formula columns (plainline only)")
            return

        for row_num in range(3, len(merged) + 3):
            r = row_num

            # priority_dtf_getting_worse: 1 if getting_worse > 0, else 0
            worksheet[f'{p_dtf_worse}{r}'] = (
                f'=IF({getting_worse}{r}>0,1,0)'
            )

            # priority_fixed_asset: region priority (0/1/2) based on section_name from REGION_PRIORITY
            section_name = merged.iloc[row_num - 3]['section_name'] if 'section_name' in merged.columns else ''
            worksheet[f'{p_fixed_asset}{r}'] = REGION_PRIORITY.get(section_name, 0)

            # priority_lx_score: weighted normalized sum, 0 if lx treatment is "No Treatment"
            worksheet[f'{p_lx_score}{r}'] = (
                f'=IF({lx_final}{r}<>"No Treatment",'
                f'({p_fixed_asset}{r}/{th("priority_lx_norm_fixed_asset")})*{th("priority_lx_weight_fixed_asset")}+'
                f'({p_tsr}{r}/{th("priority_lx_norm_tsr")})*{th("priority_lx_weight_tsr")}+'
                f'({p_dtf}{r}/{th("priority_lx_norm_dtf")})*{th("priority_lx_weight_dtf")}+'
                f'({p_dtf_worse}{r}/{th("priority_lx_norm_dtf_worse")})*{th("priority_lx_weight_dtf_worse")}+'
                f'({p_speed}{r}/{th("priority_lx_norm_speed")})*{th("priority_lx_weight_speed")},'
                f'0)'
            )

            # priority_irj_score: weighted normalized sum, 0 if irj treatment is "No Treatment"
            worksheet[f'{p_irj_score}{r}'] = (
                f'=IF({irj_final}{r}<>"No Treatment",'
                f'({p_fixed_asset}{r}/{th("priority_lx_norm_fixed_asset")})*{th("priority_lx_weight_fixed_asset")}+'
                f'({p_tsr}{r}/{th("priority_lx_norm_tsr")})*{th("priority_lx_weight_tsr")}+'
                f'({p_dtf}{r}/{th("priority_lx_norm_dtf")})*{th("priority_lx_weight_dtf")}+'
                f'({p_dtf_worse}{r}/{th("priority_lx_norm_dtf_worse")})*{th("priority_lx_weight_dtf_worse")}+'
                f'({p_speed}{r}/{th("priority_lx_norm_speed")})*{th("priority_lx_weight_speed")},'
                f'0)'
            )

            # priority_turnout_score: weighted normalized sum, 0 if turnout treatment is "No Treatment"
            worksheet[f'{p_turnout_score}{r}'] = (
                f'=IF({turnout_final}{r}<>"No Treatment",'
                f'({p_fixed_asset}{r}/{th("priority_lx_norm_fixed_asset")})*{th("priority_lx_weight_fixed_asset")}+'
                f'({p_tsr}{r}/{th("priority_lx_norm_tsr")})*{th("priority_lx_weight_tsr")}+'
                f'({p_dtf}{r}/{th("priority_lx_norm_dtf")})*{th("priority_lx_weight_dtf")}+'
                f'({p_dtf_worse}{r}/{th("priority_lx_norm_dtf_worse")})*{th("priority_lx_weight_dtf_worse")}+'
                f'({p_speed}{r}/{th("priority_lx_norm_speed")})*{th("priority_lx_weight_speed")},'
                f'0)'
            )

            # priority_bridge_score: weighted normalized sum, 0 if bridge treatment is "No Treatment"
            worksheet[f'{p_bridge_score}{r}'] = (
                f'=IF({bridge_final}{r}<>"No Treatment",'
                f'({p_fixed_asset}{r}/{th("priority_lx_norm_fixed_asset")})*{th("priority_lx_weight_fixed_asset")}+'
                f'({p_tsr}{r}/{th("priority_lx_norm_tsr")})*{th("priority_lx_weight_tsr")}+'
                f'({p_dtf}{r}/{th("priority_lx_norm_dtf")})*{th("priority_lx_weight_dtf")}+'
                f'({p_dtf_worse}{r}/{th("priority_lx_norm_dtf_worse")})*{th("priority_lx_weight_dtf_worse")}+'
                f'({p_speed}{r}/{th("priority_lx_norm_speed")})*{th("priority_lx_weight_speed")},'
                f'0)'
            )

            # priority_fixed_asset_max: max of lx, irj, turnout, bridge scores
            worksheet[f'{p_fixed_asset_max}{r}'] = (
                f'=MAX({p_lx_score}{r},{p_irj_score}{r},{p_turnout_score}{r},{p_bridge_score}{r})'
            )

        logging.info("Added priority formula columns")

    def write_tqi_trend_tab(self, writer: pd.ExcelWriter):
        """Write TQI Trend Data and TQI Trend Chart tabs."""
        from openpyxl.chart import ScatterChart, Reference, Series
        from openpyxl.worksheet.datavalidation import DataValidation

        tqi_all = self.export_data.get("agg_tqi_all")
        if tqi_all is None or tqi_all.empty:
            logging.warning("No TQI history data — skipping trend tabs")
            return

        # --- Sheet 1: TQI Trend Data (all history, flat) ---
        data_sheet_name = "TQI Trend Data"
        tqi_all.to_excel(writer, sheet_name=data_sheet_name, index=False)
        ws_data = writer.sheets[data_sheet_name]
        data_last_row = len(tqi_all) + 1  # row 1 = headers, rows 2..N+1 = data

        for col in range(1, len(tqi_all.columns) + 1):
            ws_data.cell(row=1, column=col).font = Font(bold=True)
        ws_data.column_dimensions['A'].width = 35
        ws_data.column_dimensions['B'].width = 18
        ws_data.freeze_panes = 'A2'
        ws_data.auto_filter.ref = f"A1:{get_column_letter(len(tqi_all.columns))}1"

        # --- Sheet 2: TQI Trend Chart (selector + helper table + chart) ---
        from openpyxl.workbook.defined_name import DefinedName

        chart_sheet_name = "TQI Trend Chart"
        wb = writer.book
        ws_chart = wb.create_sheet(chart_sheet_name)

        unique_dates = sorted(tqi_all['collection_date'].dropna().unique())

        # Derive line code from chainage_id prefix (e.g. CHAIN-MLX-... → ML)
        prefix_to_line = {'TLX': 'TL', 'MLX': 'ML', 'SML': 'SL', 'EML': 'EL'}
        def _line_code(cid):
            parts = cid.split('-')
            prefix = parts[1] if len(parts) >= 2 else ''
            return prefix_to_line.get(prefix, prefix)

        tqi_ids = tqi_all[['chainage_id']].drop_duplicates().copy()
        tqi_ids['line_code'] = tqi_ids['chainage_id'].apply(_line_code)
        line_ids = {
            lc: sorted(grp['chainage_id'].tolist())
            for lc, grp in tqi_ids.groupby('line_code')
        }
        line_codes_sorted = sorted(line_ids.keys())

        # Hidden column F: line code list for B1 dropdown
        for i, lc in enumerate(line_codes_sorted):
            ws_chart.cell(row=i + 1, column=6, value=lc)
        lc_range = f"'{chart_sheet_name}'!$F$1:$F${len(line_codes_sorted)}"
        ws_chart.column_dimensions['F'].hidden = True

        # Hidden columns G onwards: one column per line code with its chainage_ids
        # Named ranges (e.g. "ML") allow INDIRECT($B$1) to work as dependent dropdown
        for col_offset, lc in enumerate(line_codes_sorted):
            col = 7 + col_offset  # G, H, I, J ...
            col_letter = get_column_letter(col)
            ids = line_ids[lc]
            for row_i, cid in enumerate(ids):
                ws_chart.cell(row=row_i + 1, column=col, value=cid)
            # Named range must exactly match the value in B1
            range_ref = f"'{chart_sheet_name}'!${col_letter}$1:${col_letter}${len(ids)}"
            wb.defined_names[lc] = DefinedName(lc, attr_text=range_ref)
            ws_chart.column_dimensions[col_letter].hidden = True

        # Row 1: line code selector
        ws_chart['A1'] = "Select Line:"
        ws_chart['A1'].font = Font(bold=True)
        ws_chart['B1'] = line_codes_sorted[0] if line_codes_sorted else ""
        dv_line = DataValidation(type="list", formula1=lc_range, allow_blank=False, showDropDown=False)
        ws_chart.add_data_validation(dv_line)
        dv_line.add(ws_chart['B1'])

        # Row 2: chainage_id selector (depends on B1 via named range + INDIRECT)
        ws_chart['A2'] = "Select Chainage ID:"
        ws_chart['A2'].font = Font(bold=True)
        first_lc = line_codes_sorted[0] if line_codes_sorted else ""
        ws_chart['B2'] = line_ids[first_lc][0] if first_lc and line_ids[first_lc] else ""
        dv_id = DataValidation(type="list", formula1="INDIRECT($B$1)", allow_blank=False, showDropDown=False)
        ws_chart.add_data_validation(dv_id)
        dv_id.add(ws_chart['B2'])

        # Row 4: helper table headers; rows 5+ date/TQI/status/trend values
        # agg_tqi column order: chainage_id(A), collection_date(B), tqi(C), status(D), trend(E), ...
        for col, label in enumerate(["Date", "TQI", "Status", "Trend"], start=1):
            cell = ws_chart.cell(row=4, column=col, value=label)
            cell.font = Font(bold=True)

        for i, raw_date in enumerate(unique_dates):
            row = i + 5
            date_val = raw_date.to_pydatetime() if hasattr(raw_date, 'to_pydatetime') else raw_date
            cell = ws_chart.cell(row=row, column=1, value=date_val)
            cell.number_format = 'YYYY-MM-DD'

            # TQI (numeric) — SUMPRODUCT directly
            tqi_formula = (
                f"=IF(SUMPRODUCT(('{data_sheet_name}'!$A$2:$A${data_last_row}=$B$2)"
                f"*('{data_sheet_name}'!$B$2:$B${data_last_row}=A{row}))=0,"
                f"\"\",SUMPRODUCT(('{data_sheet_name}'!$A$2:$A${data_last_row}=$B$2)"
                f"*('{data_sheet_name}'!$B$2:$B${data_last_row}=A{row})"
                f"*('{data_sheet_name}'!$C$2:$C${data_last_row})))"
            )
            ws_chart.cell(row=row, column=2, value=tqi_formula)

            # Status and Trend (text) — use SUMPRODUCT(*ROW())-1 to get row index, then INDEX
            # SUMPRODUCT returns the absolute row number of the matched row; subtract 1 to
            # convert to a 1-based position within the range (which starts at row 2).
            row_idx = (
                f"SUMPRODUCT(('{data_sheet_name}'!$A$2:$A${data_last_row}=$B$2)"
                f"*('{data_sheet_name}'!$B$2:$B${data_last_row}=A{row})"
                f"*ROW('{data_sheet_name}'!$A$2:$A${data_last_row}))-1"
            )
            for col, data_col in [(3, "D"), (4, "E")]:  # status=D, trend=E
                formula = (
                    f"=IFERROR(INDEX('{data_sheet_name}'!${data_col}$2:${data_col}${data_last_row},"
                    f"{row_idx}),\"\")"
                )
                ws_chart.cell(row=row, column=col, value=formula)

        # XY Scatter chart
        last_helper_row = len(unique_dates) + 4
        chart = ScatterChart()
        chart.scatterStyle = "marker"
        chart.title = "TQI Trend  —  select Line in B1, Chainage ID in B2"
        chart.y_axis.title = "TQI"
        chart.y_axis.delete = False
        chart.x_axis.title = "Date"
        chart.x_axis.delete = False
        chart.x_axis.numFmt = "yyyy-mm"
        chart.x_axis.sourceLinked = False
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.majorTickMark = "out"
        chart.width = 24
        chart.height = 14

        xvalues = Reference(ws_chart, min_col=1, min_row=5, max_row=last_helper_row)
        yvalues = Reference(ws_chart, min_col=2, min_row=5, max_row=last_helper_row)
        series = Series(yvalues, xvalues)
        series.smooth = False
        chart.series.append(series)

        ws_chart.add_chart(chart, "D5")

        ws_chart.column_dimensions['A'].width = 14
        ws_chart.column_dimensions['B'].width = 35
        ws_chart.freeze_panes = 'A5'

        logging.info(f"TQI trend tabs written ({len(line_codes_sorted)} lines, {len(tqi_ids)} chainage IDs, {len(unique_dates)} dates)")

    def write_excel_report(self):
        """
        Generate an Excel workbook with one tab per line_code, joining all aggregated tables.

        """
        logging.info("Generating Excel report...")
        
        from datetime import date
        output_file = os.path.join(self.output_path, f"fmg_rail_report_{date.today().strftime('%Y%m%d')}.xlsx")
        
        # Create Excel writer
        all_line_segments = self.export_data["segments"]

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write summary tabs first
            self.write_data_sources_tab(writer)
            self.write_thresholds_tab(writer)
            self.write_tqi_trend_tab(writer)
            for line_code in self.line_codes:
                logging.info(f"Processing line_code: {line_code}")

                # Filter segments for this line_code
                line_segments = all_line_segments[all_line_segments['line_code'] == line_code].copy()

                if line_segments.empty:
                    logging.warning(f"No segments found for line_code: {line_code}")
                    continue

                # Track column sources for header row
                segment_cols = list(line_segments.columns)

                # Merge all aggregated data on chainage_id
                merged = line_segments.copy()

                # Left join each agg table and track new columns
                asset_cols = []
                if not self.export_data["agg_asset"].empty:
                    before_cols = set(merged.columns)
                    merged = merged.merge(
                        self.export_data["agg_asset"],
                        on='chainage_id',
                        how='left',
                        suffixes=('', '_asset')
                    )
                    asset_cols = [col for col in merged.columns if col not in before_cols]

                gbfi_cols = []
                if not self.export_data["agg_gbfi"].empty:
                    before_cols = set(merged.columns)
                    merged = merged.merge(
                        self.export_data["agg_gbfi"],
                        on='chainage_id',
                        how='left',
                        suffixes=('', '_gbfi')
                    )
                    gbfi_cols = [col for col in merged.columns if col not in before_cols]

                ballast_cols = []
                if not self.export_data["agg_ballast"].empty:
                    before_cols = set(merged.columns)
                    merged = merged.merge(
                        self.export_data["agg_ballast"],
                        on='chainage_id',
                        how='left',
                        suffixes=('', '_ballast')
                    )
                    ballast_cols = [col for col in merged.columns if col not in before_cols]

                moisture_cols = []
                if not self.export_data["agg_moisture"].empty:
                    before_cols = set(merged.columns)
                    merged = merged.merge(
                        self.export_data["agg_moisture"],
                        on='chainage_id',
                        how='left',
                        suffixes=('', '_moisture')
                    )
                    moisture_cols = [col for col in merged.columns if col not in before_cols]

                tsr_cols = []
                if not self.export_data["agg_tsr"].empty:
                    before_cols = set(merged.columns)
                    merged = merged.merge(
                        self.export_data["agg_tsr"],
                        on='chainage_id',
                        how='left',
                        suffixes=('', '_tsr')
                    )
                    tsr_cols = [col for col in merged.columns if col not in before_cols]

                tqi_cols = []
                if not self.export_data["agg_tqi"].empty:
                    before_cols = set(merged.columns)
                    merged = merged.merge(
                        self.export_data["agg_tqi"],
                        on='chainage_id',
                        how='left',
                        suffixes=('', '_tqi')
                    )
                    tqi_cols = [col for col in merged.columns if col not in before_cols]

                # Ensure trend columns are always present (may be absent if TQI data is empty)
                for _col in ['trend', 'step_change_type', 'step_change_date', 'trend_segment_start']:
                    if _col not in merged.columns:
                        merged[_col] = None
                        tqi_cols.append(_col)
                
                dtr_cols = []
                if not self.export_data["agg_dtr"].empty:
                    before_cols = set(merged.columns)
                    merged = merged.merge(
                        self.export_data["agg_dtr"],
                        on='chainage_id',
                        how='left',
                        suffixes=('', '_dtr')
                    )
                    dtr_cols = [col for col in merged.columns if col not in before_cols]

                recommendation_cols = []
                plainline_priority_col_names = {'priority_tsr', 'priority_dtf', 'priority_tqi_status',
                                               'priority_tqi_trend', 'priority_speed', 'priority_curve', 'priority_score'}
                fixed_asset_priority_col_names = {'priority_dtf_getting_worse', 'priority_fixed_asset', 'priority_lx_score',
                                                   'priority_irj_score', 'priority_turnout_score', 'priority_bridge_score',
                                                   'priority_fixed_asset_max'}
                if "recommendations" in self.export_data and not self.export_data["recommendations"].empty:
                    before_cols = set(merged.columns)
                    merged = merged.merge(
                        self.export_data["recommendations"],
                        on='chainage_id',
                        how='left',
                        suffixes=('', '_rec')
                    )
                    all_priority_col_names = plainline_priority_col_names | fixed_asset_priority_col_names
                    new_cols = [col for col in merged.columns if col not in before_cols]
                    recommendation_cols = [col for col in new_cols if col not in all_priority_col_names]
                    plainline_priority_cols = [col for col in new_cols if col in plainline_priority_col_names]
                    fixed_asset_priority_cols = [col for col in new_cols if col in fixed_asset_priority_col_names]

                # Sort by chainage_start_km
                merged = merged.sort_values('chainage_start_km')

                # Write to Excel tab (sheet name limited to 31 chars)
                sheet_name = f"{line_code}"[:31]
                merged.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

                # Access the worksheet to add merged header row
                worksheet = writer.sheets[sheet_name]

                # Build column group mapping
                col_groups = []

                # Determine source for each column
                start_col = 1  # Excel columns are 1-indexed

                if segment_cols:
                    end_col = start_col + len(segment_cols) - 1
                    col_groups.append(('Rail Segments', start_col, end_col))
                    start_col = end_col + 1

                if asset_cols:
                    end_col = start_col + len(asset_cols) - 1
                    col_groups.append(('Assets', start_col, end_col))
                    start_col = end_col + 1

                if gbfi_cols:
                    end_col = start_col + len(gbfi_cols) - 1
                    col_groups.append(('Ground Ballast Fouling Index (GBFI)', start_col, end_col))
                    start_col = end_col + 1

                if ballast_cols:
                    end_col = start_col + len(ballast_cols) - 1
                    col_groups.append(('Ballast', start_col, end_col))
                    start_col = end_col + 1

                if moisture_cols:
                    end_col = start_col + len(moisture_cols) - 1
                    col_groups.append(('Moisture', start_col, end_col))
                    start_col = end_col + 1

                if tsr_cols:
                    end_col = start_col + len(tsr_cols) - 1
                    col_groups.append(('Temporary Speed Restriction (TSR)', start_col, end_col))
                    start_col = end_col + 1

                if tqi_cols:
                    end_col = start_col + len(tqi_cols) - 1
                    col_groups.append(('Track Quality Indicator (TQI)', start_col, end_col))
                    start_col = end_col + 1
                    
                if dtr_cols:
                    end_col = start_col + len(dtr_cols) - 1
                    col_groups.append(('Dynamic Track Force (DTR)', start_col, end_col))
                    start_col = end_col + 1

                if recommendation_cols:
                    end_col = start_col + len(recommendation_cols) - 1
                    col_groups.append(('Recommendations', start_col, end_col))
                    start_col = end_col + 1

                if plainline_priority_cols:
                    end_col = start_col + len(plainline_priority_cols) - 1
                    col_groups.append(('Plainline Prioritisation', start_col, end_col))
                    start_col = end_col + 1

                if fixed_asset_priority_cols:
                    end_col = start_col + len(fixed_asset_priority_cols) - 1
                    col_groups.append(('Fixed Asset Prioritisation', start_col, end_col))

                add_merged_header_row(worksheet, col_groups)

                # Add Excel formulas to treatment final columns
                self.cols_list = list(merged.columns)

                # Helper function to add formulas for a treatment type.
                # If window > 0 and treatment_hierarchy is provided, the formula picks the most
                # aggressive treatment seen across +/- window rows (smoothing over ~500m of track).
                def add_treatment_formulas(model_col, override_col, final_col, window=0, treatment_hierarchy=None):
                    if model_col in merged.columns and override_col in merged.columns and final_col in merged.columns:
                        # Find column indices (1-indexed for Excel)
                        model_col_idx = self.cols_list.index(model_col) + 1
                        override_col_idx = self.cols_list.index(override_col) + 1
                        final_col_idx = self.cols_list.index(final_col) + 1

                        # Convert to Excel column letters
                        model_col_letter = get_column_letter(model_col_idx)
                        override_col_letter = get_column_letter(override_col_idx)
                        final_col_letter = get_column_letter(final_col_idx)

                        first_row = 3  # row 1 = merged header, row 2 = column names
                        last_row = len(merged) + 2

                        # Write formula to each row
                        for row_num in range(first_row, last_row + 1):
                            if window > 0 and treatment_hierarchy:
                                # Clamp window to data bounds
                                start_row = max(first_row, row_num - window)
                                end_row = min(last_row, row_num + window)
                                range_str = f'{model_col_letter}{start_row}:{model_col_letter}{end_row}'

                                # Build nested IF using COUNTIF, most aggressive first
                                # Default to the last (least aggressive) treatment
                                inner = f'"{treatment_hierarchy[-1]}"'
                                for treatment in reversed(treatment_hierarchy[:-1]):
                                    inner = f'IF(COUNTIF({range_str},"{treatment}")>0,"{treatment}",{inner})'

                                formula = f'=IF(ISBLANK({override_col_letter}{row_num}),{inner},{override_col_letter}{row_num})'
                            else:
                                formula = f'=IF(ISBLANK({override_col_letter}{row_num}),{model_col_letter}{row_num},{override_col_letter}{row_num})'

                            worksheet[f'{final_col_letter}{row_num}'] = formula

                PLAINLINE_TREATMENT_HIERARCHY = [
                    "Local Formation Renewal (Bog Hole)",
                    "Ballast Clean",
                    "Lift & Tamp",
                    "Tamp Only",
                    "No Treatment",
                ]

                # Add formulas for plainline treatment (windowed: picks most aggressive in +/- 2 segments)
                add_treatment_formulas('plainline_treatment_model', 'plainline_treatment_override', 'plainline_treatment_final',
                                       window=2, treatment_hierarchy=PLAINLINE_TREATMENT_HIERARCHY)

                # Add formulas for lx treatment
                add_treatment_formulas('lx_treatment_model', 'lx_treatment_override', 'lx_treatment_final')
                
                # Add formulas for irj treatment
                add_treatment_formulas('irj_treatment_model', 'irj_treatment_override', 'irj_treatment_final')
                
                # Add formulas for turnout treatment
                add_treatment_formulas('turnout_treatment_model', 'turnout_treatment_override', 'turnout_treatment_final')
                
                # Add formulas for bridge treatment
                add_treatment_formulas('bridge_treatment_model', 'bridge_treatment_override', 'bridge_treatment_final')

                # TEMP: Add formula columns that reference Thresholds tab for troubleshooting
                self._add_treatment_formula_columns(worksheet, merged)

                # Add formulas for problem_zone, problem_zone_group, problem_zone_length_m
                if all(col in merged.columns for col in ['plainline_treatment_final', 'lx_treatment_final', 'irj_treatment_final', 'turnout_treatment_final', 'bridge_treatment_final', 'problem_zone', 'problem_zone_group', 'problem_zone_length_m']):
                    plainline_idx = self.cols_list.index('plainline_treatment_final') + 1
                    lx_idx = self.cols_list.index('lx_treatment_final') + 1
                    irj_idx = self.cols_list.index('irj_treatment_final') + 1
                    turnout_idx = self.cols_list.index('turnout_treatment_final') + 1
                    bridge_idx = self.cols_list.index('bridge_treatment_final') + 1
                    problem_zone_idx = self.cols_list.index('problem_zone') + 1
                    problem_zone_group_idx = self.cols_list.index('problem_zone_group') + 1
                    problem_zone_length_idx = self.cols_list.index('problem_zone_length_m') + 1

                    # Get column letters
                    plainline_letter = get_column_letter(plainline_idx)
                    lx_letter = get_column_letter(lx_idx)
                    irj_letter = get_column_letter(irj_idx)
                    turnout_letter = get_column_letter(turnout_idx)
                    bridge_letter = get_column_letter(bridge_idx)
                    problem_zone_letter = get_column_letter(problem_zone_idx)
                    problem_zone_group_letter = get_column_letter(problem_zone_group_idx)
                    problem_zone_length_letter = get_column_letter(problem_zone_length_idx)

                    # Get chainage length columns if available
                    chainage_start_letter = None
                    chainage_end_letter = None
                    if 'chainage_start_km' in merged.columns and 'chainage_end_km' in merged.columns:
                        chainage_start_idx = self.cols_list.index('chainage_start_km') + 1
                        chainage_end_idx = self.cols_list.index('chainage_end_km') + 1
                        chainage_start_letter = get_column_letter(chainage_start_idx)
                        chainage_end_letter = get_column_letter(chainage_end_idx)

                    for row_num in range(3, len(merged) + 3):
                        # Formula for problem_zone: 1 if any final treatment is not "No Treatment", 0 otherwise
                        problem_zone_formula = f'=IF(OR({plainline_letter}{row_num}<>"No Treatment",{lx_letter}{row_num}<>"No Treatment",{irj_letter}{row_num}<>"No Treatment",{turnout_letter}{row_num}<>"No Treatment",{bridge_letter}{row_num}<>"No Treatment"),1,0)'
                        worksheet[f'{problem_zone_letter}{row_num}'] = problem_zone_formula

                        # Formula for problem_zone_group: consecutive problem zones get same group number
                        if row_num == 3:  # First data row
                            problem_zone_group_formula = f'=IF({problem_zone_letter}{row_num}=0,"",1)'
                        else:
                            prev_row = row_num - 1
                            problem_zone_group_formula = f'=IF({problem_zone_letter}{row_num}=0,"",IF({problem_zone_letter}{prev_row}=1,{problem_zone_group_letter}{prev_row},IF(COUNTA(${problem_zone_group_letter}$3:{problem_zone_group_letter}{prev_row})=0,1,MAX(${problem_zone_group_letter}$3:{problem_zone_group_letter}{prev_row})+1)))'
                        worksheet[f'{problem_zone_group_letter}{row_num}'] = problem_zone_group_formula

                        # Formula for problem_zone_length_m: total chainage length for the group
                        if chainage_start_letter and chainage_end_letter:
                            # Use SUMPRODUCT to calculate segment lengths and sum for the group
                            last_row = len(merged) + 2
                            problem_zone_length_formula = f'=IF({problem_zone_letter}{row_num}=0,"",SUMPRODUCT((${problem_zone_group_letter}$3:${problem_zone_group_letter}${last_row}={problem_zone_group_letter}{row_num})*(${chainage_end_letter}$3:${chainage_end_letter}${last_row}-${chainage_start_letter}$3:${chainage_start_letter}${last_row})*1000))'
                            worksheet[f'{problem_zone_length_letter}{row_num}'] = problem_zone_length_formula

                # Add priority scoring columns
                self._add_priority_columns(worksheet, merged)

                # Add autofilter on row 2 (column headers)
                last_col_letter = get_column_letter(len(merged.columns))
                last_row = len(merged) + 2  # +2 for header rows
                worksheet.auto_filter.ref = f"A2:{last_col_letter}{last_row}"

                # Set column A width wider (for chainage_id)
                worksheet.column_dimensions['A'].width = 35

                # Freeze panes at B3 (freezes row 1-2 and column A)
                worksheet.freeze_panes = 'B3'

                logging.info(f"  Written {len(merged)} rows to sheet '{sheet_name}'")

        logging.info(f"Excel report saved to: {self.output_path}")

        self.write_tqi_html_report(output_file)

    def write_tqi_html_report(self, excel_path: str):
        """Generate a self-contained HTML TQI trend report alongside the Excel file."""
        import json

        tqi_all = self.export_data.get("agg_tqi_all")
        if tqi_all is None or tqi_all.empty:
            logging.warning("No TQI history data — skipping HTML report")
            return

        def _line(cid: str) -> str:
            parts = str(cid).split("-")
            return parts[1] if len(parts) >= 2 else "?"

        df = tqi_all.copy()
        df["collection_date"] = pd.to_datetime(df["collection_date"]).dt.strftime("%Y-%m-%d")
        for _dc in ["step_change_date", "trend_segment_start"]:
            if _dc in df.columns:
                df[_dc] = df[_dc].apply(lambda x: x.strftime("%Y-%m-%d") if pd.notna(x) else None)
        df["line_code"] = df["chainage_id"].apply(_line)

        line_ids = {
            str(lc): sorted(grp["chainage_id"].dropna().unique().tolist())
            for lc, grp in df.groupby("line_code")
        }
        base_cols = ["chainage_id", "line_code", "collection_date",
                     "tqi", "status", "trend", "trend_slope", "trend_r_squared",
                     "step_change_type", "step_change_date", "trend_segment_start"]
        coord_cols = [c for c in ["mid_coord_lat", "mid_coord_lng"] if c in df.columns]
        records = (
            df[base_cols + coord_cols]
            .where(df.notna(), None)
            .to_dict("records")
        )

        data_json    = json.dumps(records,  ensure_ascii=False)
        line_id_json = json.dumps(line_ids, ensure_ascii=False)

        # Re-use the same HTML template from generate_tqi_html.py
        from rail_dog.utils.html_utils import tqi_trend_html
        html = tqi_trend_html(data_json, line_id_json, mapbox_token=self.mapbox_token)

        out_path = str(excel_path).replace(".xlsx", "_tqi_trend.html")
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(html)
        logging.info(f"TQI HTML report saved to: {out_path}")

def add_merged_header_row(worksheet, col_groups: list):
    # Insert merged cells in first row

    for source_name, start_idx, end_idx in col_groups:
        start_letter = get_column_letter(start_idx)
        end_letter = get_column_letter(end_idx)

        # Merge cells
        if start_idx == end_idx:
            cell_ref = f"{start_letter}1"
        else:
            cell_ref = f"{start_letter}1:{end_letter}1"
            worksheet.merge_cells(cell_ref)

        # Write source label
        cell = worksheet[f"{start_letter}1"]
        cell.value = source_name
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add background color based on source
        colors = {
            'Rail Segments': 'DDEBF7',  # Light blue
            'Assets': 'E2EFDA',    # Light green
            'Ground Ballast Fouling Index (GBFI)': 'FCE4D6',     # Light orange
            'Ballast': 'F8CBAD',    # Light peach
            'Temporary Speed Restriction (TSR)': 'FFF2CC',      # Light yellow
            'Track Quality Indicator (TQI)': 'F4B084',      # Light coral
            'Dynamic Track Force (DTR)': 'E4DFEC',      # Light purple
            'Recommendations': 'D5F4E6',      # Light mint green
            'Plainline Prioritisation': 'FFD6E8',   # Light pink
            'Fixed Asset Prioritisation': 'FADBD8', # Light salmon
        }
        if source_name in colors:
            cell.fill = PatternFill(start_color=colors[source_name],
                                end_color=colors[source_name],
                                fill_type='solid')

    # Set row height for header
    worksheet.row_dimensions[1].height = 20

